# Main.py

import sys
import os
import json
import subprocess, re
import logging
import shutil
import ctypes
import tempfile
from ctypes import wintypes
from pathlib import Path

from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtCore import QUrl
from PyQt5.QtGui import QDesktopServices
from PyQt5.QtWidgets import QApplication, QMessageBox

from AutoTest import AutoTest_Start
from Main_Window import Ui_MainWindow
from device_lock import DeviceLock, describe_holder
from multiprocessing import Process, Queue, freeze_support

from request_API import (
    login_step1,
    get_curriculum_response,
    class_list,
    student_list_by_class,
    authenticate_study_access,
    authenticate_study_access_detailed,
    get_study_access_auth,
    get_parent_report,
    post_attendance_curriculum,
    get_witti_school_main,
    get_witti_app_main,
    get_aram_bookworld_subject,
    get_witti_school_ebook_main,
    get_tv_main,
    get_teacher_activity_report,
)

# PyInstaller windowed 빌드(console=False)에서 adb.exe/scrcpy.exe 같은 콘솔 프로그램을
# 실행하면 검은 콘솔 창이 순간적으로 떴다 사라진다. subprocess 기본값에
# CREATE_NO_WINDOW를 주입해 앱/airtest 내부 호출까지 모두 창 없이 실행한다.
if sys.platform == "win32":
    _CREATE_NO_WINDOW = 0x08000000
    _orig_popen_init = subprocess.Popen.__init__

    def _popen_init_no_window(self, *args, **kwargs):
        if not kwargs.get("creationflags"):
            kwargs["creationflags"] = _CREATE_NO_WINDOW
        _orig_popen_init(self, *args, **kwargs)

    subprocess.Popen.__init__ = _popen_init_no_window


def load_local_config():
    """
    gitignore된 local_config.json에서 로컬 전용 설정(테스트 계정 등)을 읽는다.
    파일이 없으면 빈 dict를 반환한다 (local_config.example.json 참고).

    탐색 순서 (앞의 것이 우선):
      1) exe(또는 스크립트)가 놓인 폴더  → 배포 후 계정을 바꿔 쓸 수 있게
      2) PyInstaller 번들 내부(_MEIPASS) → 빌드 시 포함된 기본값
    onefile 빌드에서 __file__은 임시 해제 폴더를 가리키므로
    __file__만 보면 exe 옆에 둔 설정 파일을 영영 못 찾는다.
    """
    candidates = []
    if getattr(sys, "frozen", False):
        candidates.append(os.path.dirname(sys.executable))
        if hasattr(sys, "_MEIPASS"):
            candidates.append(sys._MEIPASS)
    else:
        candidates.append(os.path.dirname(os.path.abspath(__file__)))

    for base in candidates:
        path = os.path.join(base, "local_config.json")
        try:
            with open(path, encoding="utf-8") as f:
                return json.load(f)
        except (OSError, ValueError):
            continue
    return {}


# ── adb 서버 포트 ─────────────────────────────────────────────────────
# 기본 포트 5037을 그대로 쓴다. 전용 포트로 옮겨봤지만 손해가 더 컸다.
# USB 기기는 adb '서버' 단위로 배타적이라, 서버가 둘이면 같은 기기를 두고
# 선착순으로 다투고 진 쪽에는 그 기기가 아예 보이지 않는다. Android Studio가
# 5037에 자기 서버를 띄우므로 우리가 나가면 서버가 둘이 된다.
# (실측: 전용 포트 1대 vs 5037 USB 2대 + 무선 2대. 자세한 경위는 커밋 478533b)
#
# 서버가 재시작되어 연결이 끊기는 것은 adb_recovery가, 같은 기기를 두 인스턴스가
# 동시에 테스트하는 것은 device_lock이 담당한다. 포트로 막을 수 있는 게 아니다.
_DEFAULT_ADB_SERVER_PORT = 5037

# 포트를 옮겨야 할 때만 쓰는 탈출구. 설정 파일이 아니라 환경변수인 이유는,
# 인스턴스마다 값이 갈리면 서버가 쪼개져 위의 USB 쟁탈이 그대로 생기기 때문이다.
# 옮길 포트가 비어 있는지 반드시 확인할 것 (5040은 svchost가 물고 있었다).
_ADB_PORT_ENV = "WITTI_ADB_SERVER_PORT"


def _instance_dir():
    """이 인스턴스의 기준 폴더 (exe가 놓인 폴더 / 개발 시 프로젝트 폴더)."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


def get_adb_port():
    """이 앱이 쓰는 adb 서버 포트. 하위 프로세스도 같은 서버를 보게 전파한다."""
    try:
        port = int(os.environ.get(_ADB_PORT_ENV, "") or _DEFAULT_ADB_SERVER_PORT)
    except ValueError:
        port = _DEFAULT_ADB_SERVER_PORT
    # scrcpy와 워커 프로세스는 각자 adb 클라이언트를 띄운다. 환경변수를 함께
    # 넘겨야 포트를 옮겼을 때 그쪽만 5037에 남는 일이 없다.
    os.environ["ANDROID_ADB_SERVER_PORT"] = str(port)
    return port


def get_adb_path():
    """
    Return bundled adb.exe path.
    In development mode, use ./adb/adb.exe.
    In PyInstaller onefile mode, files are unpacked under sys._MEIPASS.
    """
    if hasattr(sys, "_MEIPASS"):
        base = sys._MEIPASS
    else:
        base = os.path.abspath(".")
    return os.path.join(base, "adb", "adb.exe")


def get_scrcpy_path():
    """
    Return bundled scrcpy.exe path.
    In development mode, use ./scrcpy/scrcpy.exe.
    In PyInstaller onefile mode, files are unpacked under sys._MEIPASS.
    Falls back to scrcpy on PATH.
    """
    if hasattr(sys, "_MEIPASS"):
        base = sys._MEIPASS
    else:
        base = os.path.abspath(".")
    candidate = os.path.join(base, "scrcpy", "scrcpy.exe")
    if os.path.exists(candidate):
        return candidate
    return shutil.which("scrcpy")


def run_adb(args, **popen_kwargs):
    """
    Run adb command and return output.
    Uses check_output by default.
    Example: run_adb(["devices", "-l"], text=True)
    """
    adb = get_adb_path()
    # 포트를 명시한다. 환경변수만 믿으면 포트를 옮겼을 때 어딘가 한 곳이
    # 기본값으로 새어도 조용히 엉뚱한 서버에 붙어 원인을 찾기 어려워진다.
    return subprocess.check_output([adb, "-P", str(get_adb_port())] + args, **popen_kwargs)


def ensure_adb_server():
    """
    이 앱이 쓸 adb 서버를 띄운다.

    kill-server는 하지 않는다. 다른 인스턴스가 테스트 중일 때 서버를 죽이면
    Wi-Fi 연결 기기가 떨어지고 미러링(scrcpy) 터널도 끊긴다.
    이미 서버가 떠 있으면 start-server는 아무 동작도 하지 않는다.

    조회 전에 반드시 불러야 한다. 서버가 없으면 adb 클라이언트가 자기가
    띄우는데, 그 기동 비용(실측 약 2초)을 UI 스레드가 그대로 뒤집어쓴다.
    Failures are ignored to keep app flow running.
    """
    port = get_adb_port()

    # check_output에 stdout을 넘기면 ValueError로 죽는다. 출력을 버리면서
    # 예외만 무시하려면 run을 써야 한다. (check_output으로 두면 서버가
    # 한 번도 뜨지 않아 이 함수가 조용히 무의미해진다)
    try:
        subprocess.run([get_adb_path(), "-P", str(port), "start-server"],
                       check=False, stdout=subprocess.DEVNULL,
                       stderr=subprocess.DEVNULL, timeout=15)
    except Exception:
        pass

    return port


def describe_adb_server():
    """현재 adb 서버 설정을 사람이 읽을 수 있는 한 줄로."""
    return f"adb 서버: 포트 {get_adb_port()} (인스턴스 폴더 {_instance_dir()})"


def ensure_adb_keys(app_name="AutoTest"):
    """
    1) Pin key path to %LOCALAPPDATA%\\{app_name}\\.android
    2) Copy existing user keys once if present
    3) Generate keys with bundled adb start-server if missing
    4) Always set ADB_VENDOR_KEYS to the pinned key folder
    """
    user_home = Path(os.path.expandvars(r"%USERPROFILE%"))
    user_dot_android = user_home / ".android"
    user_key = user_dot_android / "adbkey"
    user_key_pub = user_dot_android / "adbkey.pub"

    stable_root = Path(os.path.expandvars(r"%LOCALAPPDATA%")) / app_name / ".android"
    stable_root.mkdir(parents=True, exist_ok=True)
    stable_key = stable_root / "adbkey"
    stable_key_pub = stable_root / "adbkey.pub"

    if user_key.exists() and not stable_key.exists():
        try:
            shutil.copy2(user_key, stable_key)
            if user_key_pub.exists():
                shutil.copy2(user_key_pub, stable_key_pub)
        except Exception:
            pass

    os.environ["ADB_VENDOR_KEYS"] = str(stable_root)

    if not stable_key.exists():
        # 키 생성용 start-server도 같은 포트로 띄운다.
        # 포트를 빼먹으면 포트를 옮겼을 때 여기서만 5037 서버가 따로 생긴다.
        adb = get_adb_path()
        try:
            subprocess.run([adb, "-P", str(get_adb_port()), "start-server"], check=False,
                           stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception:
            pass


ensure_adb_keys(app_name="AutoTest")


def worker_main(
    log_queue,
    btn_name,
    device_name,
    device_label,
    inputId,
    inputPwd,
    subjCd,
    itemCd,
    curtnSeq,
    title_name,
    server,
    study_access_mem_nm,
    study_access_mem_id,
    study_access_auth_token,
    selected_class_id="",
):
    import builtins, traceback
    def _print_via_queue(*args, sep=" ", end="\n", **kwargs):
        msg = sep.join(str(a) for a in args) + end
        log_queue.put(msg.rstrip("\n"))
    builtins.print = _print_via_queue

    try:
        print("워커 프로세스 시작")
        AutoTest_Start(
            btn_name,
            device_name,
            device_label,
            inputId,
            inputPwd,
            subjCd,
            itemCd,
            curtnSeq,
            title_name,
            server,
            study_access_mem_nm,
            study_access_mem_id,
            study_access_auth_token,
            selected_class_id,
        )
    except Exception as e:
        print(f"[ERROR] AutoTest_Start 중 예외: {e!r}")
        print(traceback.format_exc())


def worker_all_api_test(log_queue, user_id, user_pwd, server, device_label, steps,
                        gui_ctx=None):
    """ALL API 테스트.
    - GUI에서 학생 선택 시: 그 1명만 테스트
    - GUI 미선택 시: 선생님(teacherMemId 있을 때) + 전체 학생 테스트
    매 실행마다 새 파일 (ALL_api_test_{YYMMDD}_{HHMMSS}.xlsx) 생성.
    파일에는 요약 시트 + 타깃별 시트(memNm) 포함.
    """
    import builtins
    import traceback
    from datetime import datetime, date
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    server_label = {"Prod": "Prod", "QA": "QA", "Dev": "Dev"}.get(server, server)

    match server:
        case "Prod":
            srv = "api"
        case "QA":
            srv = "qa-api"
        case "Dev":
            srv = "dev-api"
        case _:
            srv = "api"

    def _print_via_queue(*args, sep=" ", end="\n", **kwargs):
        msg = sep.join(str(a) for a in args) + end
        log_queue.put(msg.rstrip("\n"))

    builtins.print = _print_via_queue

    # 타깃별 누적 저장소
    results_by_target = {}  # {sheet_label: [(api_name, method, path, status, mark, rd, ed), ...]}
    target_meta = []  # [{"label", "type", "mem_id", "mem_nm"}, ...]

    def _make_recorder(target_results):
        def _record(api_name, method, path, resp=None, error=None):
            if resp is not None:
                status = resp.status_code
                ok = resp.ok
                result_data = resp.text[:1000] if ok else ""
                error_detail = "" if ok else resp.text[:1000]
            elif error is not None:
                status = 0
                ok = False
                result_data = ""
                error_detail = str(error)
            else:
                status = 0
                ok = False
                result_data = ""
                error_detail = "unknown"
            mark = "PASS" if ok else "FAIL"
            target_results.append((api_name, method, path, status, mark, result_data, error_detail))
            print(f"  [{mark}] {method:6s} {path} -> {status} {error_detail[:120]}")
        return _record

    def _register_target(base_label, type_, mem_id, mem_nm, target_results,
                          class_id="", class_nm="",
                          tested_at="", student_nm="", student_id="",
                          login_id="", target_age=""):
        """시트명 충돌(동명이인) 방지 후 누적 저장소에 등록."""
        sheet_label = base_label or "unnamed"
        n = 2
        existing = {m["label"] for m in target_meta}
        while sheet_label in existing:
            sheet_label = f"{base_label}_{n}"
            n += 1
        target_meta.append({
            "label": sheet_label,
            "type": type_,
            "mem_id": mem_id,
            "mem_nm": mem_nm,
            "class_id": class_id,
            "class_nm": class_nm,
            "tested_at": tested_at,
            "student_nm": student_nm,
            "student_id": student_id,
            "login_id": login_id,
            "target_age": target_age,
        })
        results_by_target[sheet_label] = target_results
        return sheet_label

    try:
        print(f"[INFO] ALL API test started ({len(steps)}개 steps)")

        # 사용자가 "수업시작"을 명시 선택했는지 (성공 기록 여부 결정 — 실패는 항상 기록)
        record_study_access_success = "수업시작 (study/access)" in steps

        # ── 로그인 & 반 정보 ──
        token = login_step1(user_id, user_pwd, srv)
        if not token:
            print("[ERROR] login failed.")
            return

        class_resp = class_list(token, user_id, srv)
        if class_resp is None:
            print("[ERROR] class list failed.")
            return

        classes = class_resp.json().get("result", {}).get("classList", [])
        if not classes:
            print("[ERROR] no classes found.")
            return

        _gui = gui_ctx or {}
        gui_student_id = str(_gui.get("student_id") or "").strip()
        gui_class_id = str(_gui.get("class_id") or "").strip()

        # ── 테스트 대상 반 결정 ──
        if gui_student_id:
            # GUI 학생 선택 시: 해당 반 1개만
            first_class = classes[0]
            classes_to_test = [{
                "classId": gui_class_id or str(first_class.get("classId", "")),
                "classNm": _gui.get("class_nm") or str(first_class.get("classNm", "")),
                "targetAge": _gui.get("target_age") or str(first_class.get("targetAge", "")),
            }]
        elif gui_class_id:
            # GUI 반 선택 시: 그 반 1개만
            classes_to_test = [{
                "classId": gui_class_id,
                "classNm": _gui.get("class_nm") or "",
                "targetAge": _gui.get("target_age") or "",
            }]
        else:
            # 미선택 시: 전체 반
            classes_to_test = [
                {
                    "classId": str(c.get("classId", "")),
                    "classNm": str(c.get("classNm", "")),
                    "targetAge": str(c.get("targetAge", "")),
                }
                for c in classes
            ]

        print(f"[INFO] 테스트 대상 반: {len(classes_to_test)}개")
        for ci_info in classes_to_test:
            print(f"  - {ci_info['classNm']} ({ci_info['classId']})")

        # ── 타깃 결정 (반별 순회) ──
        targets = []
        all_class_names = []

        for ci_info in classes_to_test:
            class_id = ci_info["classId"]
            class_nm = ci_info["classNm"]
            target_age = ci_info["targetAge"]
            all_class_names.append(f"{class_nm}({class_id})")
            print(f"\n[INFO] class: {class_nm} ({class_id})")

            if gui_student_id:
                # GUI 학생 선택 시: 1명만
                targets.append({
                    "type": "선택",
                    "studentId": gui_student_id,
                    "studentNm": _gui.get("student_nm", "") or "선택",
                    "loginId": user_id,
                    "preset_token": _gui.get("auth_token") or "",
                    "preset_child_id": _gui.get("mem_id") or "",
                    "class_id": class_id,
                    "class_nm": class_nm,
                    "target_age": target_age,
                })
                break  # 학생 선택 시 반 루프 불필요

            stu_resp = student_list_by_class(token, class_id, srv)
            if stu_resp is None:
                print(f"[WARN] student list failed for class {class_nm}, skip")
                continue

            stu_result = stu_resp.json().get("result", {})
            students_raw = stu_result.get("studentList", [])
            teacher_mem_id = str(stu_result.get("teacherMemId", "")).strip()
            teacher_mem_nm = str(
                stu_result.get("teacherMemNm")
                or stu_result.get("teacherNm")
                or stu_result.get("memNm")
                or ""
            ).strip()

            if teacher_mem_id:
                targets.append({
                    "type": "선생님",
                    "studentId": teacher_mem_id,
                    "studentNm": teacher_mem_nm or "선생님",
                    "loginId": user_id,
                    "class_id": class_id,
                    "class_nm": class_nm,
                    "target_age": target_age,
                })
            sorted_students = sorted(
                students_raw,
                key=lambda s: str(s.get("studentNm", "")).strip(),
            )
            for s in sorted_students:
                sid = str(s.get("studentId", "")).strip()
                if not sid:
                    continue
                lid = str(
                    s.get("loginId")
                    or s.get("studentLoginId")
                    or user_id
                ).strip()
                targets.append({
                    "type": "학생",
                    "studentId": sid,
                    "studentNm": str(s.get("studentNm", "")).strip() or "-",
                    "loginId": lid,
                    "class_id": class_id,
                    "class_nm": class_nm,
                    "target_age": target_age,
                })

        class_summary = ", ".join(all_class_names) if len(all_class_names) <= 3 else f"{len(all_class_names)}개 반"

        if not targets:
            print("[ERROR] no targets to test")
            return

        n_teacher = sum(1 for t in targets if t["type"] == "선생님")
        n_student = sum(1 for t in targets if t["type"] in ("학생", "선택"))
        print(f"[INFO] 총 {len(targets)}개 타깃 (선생님: {n_teacher}, 학생/선택: {n_student})")

        # ── 타깃별 실행 루프 ──
        for ti, tgt in enumerate(targets, 1):
            print(f"\n{'='*60}")
            print(f"[{ti}/{len(targets)}] {tgt['type']}: {tgt['studentNm']} (studentId={tgt['studentId']}) | 반: {tgt.get('class_nm', '')}")
            print(f"{'='*60}")

            target_tested_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            target_results = []
            _record = _make_recorder(target_results)
            ctx = {"prod_id": "P0001", "ptnr_id": "1102"}

            # study/access (GUI preset 있으면 재사용, 없으면 호출)
            child_token = ""
            child_id = ""
            mem_nm_resp = ""
            access_child_age = None
            if tgt.get("preset_token") and tgt.get("preset_child_id"):
                child_token = tgt["preset_token"]
                child_id = tgt["preset_child_id"]
                mem_nm_resp = tgt["studentNm"]
                if record_study_access_success:
                    _record("수업시작", "POST", "/authenticate/study/access",
                            resp=type("R", (), {"status_code": 200, "ok": True, "text": "(GUI preset)"})())
            else:
                try:
                    access_result = authenticate_study_access_detailed(
                        tgt["studentId"], tgt["loginId"], srv,
                        access_type="T" if tgt["type"] == "선생님" else "C",
                    )
                    access_data = access_result.get("data") if isinstance(access_result, dict) else None
                    if not isinstance(access_data, dict):
                        access_data = {}
                    api_res = access_data.get("result", {}) if isinstance(access_data.get("result", {}), dict) else {}

                    child_token = str(api_res.get("authToken", "")).strip()
                    child_id = str(api_res.get("memId", "")).strip()
                    mem_nm_resp = str(api_res.get("memNm", "")).strip() or tgt["studentNm"]

                    # childList에서 해당 계정의 childAge 추출
                    child_list = api_res.get("childList") or []
                    access_child_age = None
                    for ch in child_list:
                        if str(ch.get("childId", "")) == child_id:
                            access_child_age = ch.get("childAge")
                            break
                    if access_child_age is None and child_list:
                        access_child_age = child_list[0].get("childAge")

                    if child_token and child_id:
                        if record_study_access_success:
                            _record("수업시작", "POST", "/authenticate/study/access",
                                    resp=type("R", (), {"status_code": 200, "ok": True, "text": str(access_data)[:1000]})())
                    else:
                        _record("수업시작", "POST", "/authenticate/study/access",
                                error="no child token/id in response")
                        # study/access 실패 → 이 타깃은 결과만 기록하고 다음 타깃
                        _register_target(tgt["studentNm"], tgt["type"],
                                         tgt["studentId"], tgt["studentNm"], target_results,
                                         tgt.get("class_id", ""), tgt.get("class_nm", ""),
                                         tested_at=target_tested_at,
                                         student_nm=tgt["studentNm"], student_id=tgt["studentId"],
                                         login_id=tgt.get("loginId", ""),
                                         target_age=tgt.get("target_age", ""))
                        continue
                except Exception as e:
                    _record("수업시작", "POST", "/authenticate/study/access", error=e)
                    _register_target(tgt["studentNm"], tgt["type"],
                                     tgt["studentId"], tgt["studentNm"], target_results,
                                     tgt.get("class_id", ""), tgt.get("class_nm", ""),
                                     tested_at=target_tested_at,
                                     student_nm=tgt["studentNm"], student_id=tgt["studentId"],
                                     login_id=tgt.get("loginId", ""),
                                     target_age=tgt.get("target_age", ""))
                    continue

            print(f"  childToken: {child_token[:20]}..., childId: {child_id}, memNm: {mem_nm_resp}")

            # report용 파라미터
            # childAge=0(늘봄)이면 curriculumTp=1, 아니면 0
            child_age_int = int(access_child_age) if access_child_age is not None else 0
            curriculum_tp = 1 if child_age_int == 0 else 0
            year, month, week = 0, 0, 0
            cur_resp = get_curriculum_response(child_token, child_id, srv)
            if cur_resp is not None and cur_resp.ok:
                cur_result = cur_resp.json().get("result", {})
                year = date.today().year
                month = cur_result.get("month", 0)
                week = cur_result.get("week", 0)
                if child_age_int != 0:
                    child_age_int = cur_result.get("childAge", child_age_int)

            # API 호출 함수 매핑 (타깃 컨텍스트에 종속 → 매 타깃마다 새 정의)
            def _call_curriculum():
                resp = get_curriculum_response(child_token, child_id, srv)
                return ("커리큘럼 조회", "GET", "/witti-box/curriculum", resp)

            def _call_attendance_curriculum():
                resp = post_attendance_curriculum(child_token, srv)
                return ("출석 시간 전송", "POST", "/witti-app/attendance/curriculum", resp)

            def _call_witti_school_main():
                resp = get_witti_school_main(child_token, srv)
                if resp is not None and resp.ok:
                    sm = resp.json().get("result", {})
                    pl = sm.get("prodList") or sm.get("productList") or []
                    if pl:
                        ctx["prod_id"] = pl[0].get("prodId", ctx["prod_id"])
                        ctx["ptnr_id"] = str(pl[0].get("ptnrId", ctx["ptnr_id"]))
                return ("위티스쿨 메인", "GET", "/witti-school/main", resp)

            def _call_witti_app_main():
                resp = get_witti_app_main(child_token, srv)
                return ("보유 위티팡 조회", "GET", "/witti-app/main", resp)

            def _call_aram_bookworld_subject():
                resp = get_aram_bookworld_subject(child_token, ctx["ptnr_id"], ctx["prod_id"], srv)
                return ("아람북월드 과목", "GET", "/witti-school/aram-bookworld/subject", resp)

            def _call_ebook_main():
                resp = get_witti_school_ebook_main(child_token, srv)
                return ("도서관 메인", "GET", "/witti-school/e-book/main", resp)

            def _call_tv_main():
                resp = get_tv_main(child_token, srv)
                return ("위티TV 메인", "GET", "/tv/main", resp)

            def _call_teacher_activity_report():
                resp = get_teacher_activity_report(child_token, child_id, child_age_int, curriculum_tp, year, month, week, srv)
                return ("선생님 활동현황", "POST", "/report/teacherActivityReport", resp)

            def _call_parent_report():
                resp = get_parent_report(child_token, child_id, child_age_int, curriculum_tp, year, month, week, srv)
                return ("학습 리포트", "POST", "/report/parentReport", resp)

            CALL_MAP = {
                "커리큘럼 조회 (curriculum)": _call_curriculum,
                "출석 시간 전송 (attendance/curriculum)": _call_attendance_curriculum,
                "위티스쿨 메인 (witti-school/main)": _call_witti_school_main,
                "보유 위티팡 조회 (witti-app/main)": _call_witti_app_main,
                "아람북월드 과목 (aram-bookworld/subject)": _call_aram_bookworld_subject,
                "도서관 메인 (e-book/main)": _call_ebook_main,
                "위티TV 메인 (tv/main)": _call_tv_main,
                "선생님 활동현황 (report/teacherActivityReport)": _call_teacher_activity_report,
                "학습 리포트 > 부모(학생) / 주간 (report/parentReport)": _call_parent_report,
            }

            for step_name in steps:
                if step_name == "수업시작 (study/access)":
                    continue
                call_fn = CALL_MAP.get(step_name)
                if call_fn is None:
                    print(f"[WARN] ALL 매핑에 없는 API: {step_name}, skip")
                    continue
                print(f"\n--- {step_name} ---")
                try:
                    display_name, method, path, resp = call_fn()
                    if resp is not None:
                        _record(display_name, method, path, resp=resp)
                    else:
                        _record(display_name, method, path, error="API returned None")
                except Exception as e:
                    _record(step_name, "?", "?", error=e)

            base_label = mem_nm_resp or tgt["studentNm"] or "unnamed"
            _register_target(base_label, tgt["type"],
                             child_id or tgt["studentId"], mem_nm_resp or tgt["studentNm"],
                             target_results,
                             tgt.get("class_id", ""), tgt.get("class_nm", ""),
                             tested_at=target_tested_at,
                             student_nm=tgt["studentNm"], student_id=tgt["studentId"],
                             login_id=tgt.get("loginId", ""),
                             target_age=tgt.get("target_age", ""))

            t_pass = sum(1 for r in target_results if r[4] == "PASS")
            t_fail = sum(1 for r in target_results if r[4] == "FAIL")
            print(f"\n[{tgt['studentNm']}] 총 {len(target_results)}개 | PASS: {t_pass} | FAIL: {t_fail}")

        # ── 엑셀 저장 (매 실행 새 파일) ──
        date_str = datetime.now().strftime("%y%m%d")
        time_str = datetime.now().strftime("%H%M%S")
        report_dir = os.path.join(os.getcwd(), "test_report", "api_test")
        os.makedirs(report_dir, exist_ok=True)
        file_path = os.path.join(report_dir, f"ALL_api_test_{date_str}_{time_str}.xlsx")

        wb = Workbook()
        wb.remove(wb.active)

        # 스타일
        header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_font = Font(name="맑은 고딕", size=10)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        cell_align = Alignment(vertical="center", wrap_text=True)
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        pass_font = Font(name="맑은 고딕", size=10, color="006100", bold=True)
        fail_font = Font(name="맑은 고딕", size=10, color="9C0006", bold=True)
        link_font = Font(name="맑은 고딕", size=10, color="0563C1", underline="single")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # 정렬: 반별 그룹 → 선생님 우선 → 학생 가나다순
        sorted_meta = sorted(
            target_meta,
            key=lambda m: (m.get("class_nm", ""), 0 if m["type"] == "선생님" else 1, m["mem_nm"]),
        )

        # ── 요약 시트 ──
        ws_sum = wb.create_sheet(title="요약")
        total_targets = len(sorted_meta)
        total_pass = sum(sum(1 for r in v if r[4] == "PASS") for v in results_by_target.values())
        total_fail = sum(sum(1 for r in v if r[4] == "FAIL") for v in results_by_target.values())
        total_apis = total_pass + total_fail

        ws_sum.merge_cells("A1:O1")
        sc = ws_sum["A1"]
        sc.value = (
            f"API 테스트 결과 — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  "
            f"서버: {server_label}  |  계정: {user_id}  |  "
            f"반: {class_summary}  |  "
            f"총 대상: {total_targets}명  |  "
            f"총 API: {total_apis}개  |  PASS: {total_pass}  |  FAIL: {total_fail}"
        )
        sc.font = Font(name="맑은 고딕", bold=True, size=12)
        sc.alignment = Alignment(vertical="center")
        ws_sum.row_dimensions[1].height = 30

        ws_sum.append([])
        sum_headers = [
            "No.", "tested_at", "반", "classId", "targetAge",
            "대상명", "studentId", "loginIdUsed", "memNm", "memId",
            "종류", "총 API", "PASS", "FAIL", "성공률",
        ]
        for ci, h in enumerate(sum_headers, 1):
            c = ws_sum.cell(row=3, column=ci, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align
            c.border = thin_border
        ws_sum.row_dimensions[3].height = 25

        # 시트명 → 31자 자른 형태 미리 계산 (하이퍼링크용)
        sheet_name_map = {}  # label → actual sheet name
        used_names = {"요약"}
        for meta in sorted_meta:
            base = meta["label"][:31]
            sname = base
            if sname in used_names:
                trim = base[:28]
                k = 2
                while f"{trim}_{k}" in used_names:
                    k += 1
                sname = f"{trim}_{k}"
            sheet_name_map[meta["label"]] = sname
            used_names.add(sname)

        for i, meta in enumerate(sorted_meta, 1):
            row = i + 3
            tr = results_by_target.get(meta["label"], [])
            t_pass = sum(1 for r in tr if r[4] == "PASS")
            t_fail = sum(1 for r in tr if r[4] == "FAIL")
            t_total = t_pass + t_fail
            rate = f"{(t_pass / t_total * 100):.1f}%" if t_total else "-"

            student_nm = meta.get("student_nm", "") or meta["mem_nm"]
            vals = [
                i,
                meta.get("tested_at", ""),
                meta.get("class_nm", ""),
                meta.get("class_id", ""),
                meta.get("target_age", ""),
                student_nm,
                meta.get("student_id", ""),
                meta.get("login_id", ""),
                meta["mem_nm"],
                meta["mem_id"],
                meta["type"],
                t_total, t_pass, t_fail, rate,
            ]
            for ci, v in enumerate(vals, 1):
                c = ws_sum.cell(row=row, column=ci, value=v)
                c.font = cell_font
                c.border = thin_border
                c.alignment = left_align if ci == 6 else center_align

            # 대상명 셀(column 6) → 해당 시트로 하이퍼링크
            link_cell = ws_sum.cell(row=row, column=6)
            link_cell.hyperlink = f"#'{sheet_name_map[meta['label']]}'!A1"
            link_cell.font = link_font

            # studentNm vs memNm 불일치 강조 (column 6 = 대상명, column 9 = memNm)
            api_mem_nm = meta["mem_nm"]
            if student_nm and api_mem_nm and student_nm != api_mem_nm:
                mismatch_fill = PatternFill(fill_type="solid", start_color="FFCCCC", end_color="FFCCCC")
                mismatch_font = Font(name="맑은 고딕", size=10, color="FF0000", bold=True)
                # 하이퍼링크 폰트 색을 살리되 배경만 mismatch
                ws_sum.cell(row=row, column=6).fill = mismatch_fill
                memnm_cell = ws_sum.cell(row=row, column=9)
                memnm_cell.fill = mismatch_fill
                memnm_cell.font = mismatch_font

            # PASS 셀(column 13) 강조 (전부 통과일 때만 녹색)
            pass_cell = ws_sum.cell(row=row, column=13)
            if t_fail == 0 and t_pass > 0:
                pass_cell.fill = pass_fill
                pass_cell.font = pass_font

            # FAIL 셀(column 14) 강조 (실패 있을 때만 빨강)
            fail_cell = ws_sum.cell(row=row, column=14)
            if t_fail > 0:
                fail_cell.fill = fail_fill
                fail_cell.font = fail_font

        # API별 실패 패턴 섹션
        api_fail_count = {}  # (method, path) → set of labels
        for label, rs in results_by_target.items():
            for (api_name, method, path, status, mark, rd, ed) in rs:
                key = (method, path)
                if key not in api_fail_count:
                    api_fail_count[key] = set()
                if mark == "FAIL":
                    api_fail_count[key].add(label)

        fail_apis = [(m, p, len(s)) for (m, p), s in api_fail_count.items() if s]
        fail_apis.sort(key=lambda x: -x[2])

        if fail_apis:
            section_row = len(sorted_meta) + 5
            ws_sum.merge_cells(start_row=section_row, start_column=1,
                               end_row=section_row, end_column=15)
            sc2 = ws_sum.cell(row=section_row, column=1,
                              value="API별 실패 패턴 (1회 이상 FAIL)")
            sc2.font = Font(name="맑은 고딕", bold=True, size=11)
            sc2.alignment = Alignment(vertical="center")
            ws_sum.row_dimensions[section_row].height = 25

            api_headers = ["Method", "API Path", "FAIL 대상 수"]
            for ci, h in enumerate(api_headers, 1):
                c = ws_sum.cell(row=section_row + 1, column=ci, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = header_align
                c.border = thin_border
            ws_sum.row_dimensions[section_row + 1].height = 25

            for i, (method, path, fail_n) in enumerate(fail_apis, 1):
                row = section_row + 1 + i
                vals = [method, path, f"{fail_n}/{total_targets}"]
                for ci, v in enumerate(vals, 1):
                    c = ws_sum.cell(row=row, column=ci, value=v)
                    c.font = cell_font
                    c.border = thin_border
                    c.alignment = left_align if ci == 2 else center_align
                # 모든 대상에서 실패 → 빨간 강조 (환경 이슈 의심)
                if fail_n == total_targets:
                    for ci in range(1, 4):
                        ws_sum.cell(row=row, column=ci).fill = fail_fill

        ws_sum.column_dimensions["A"].width = 6   # No.
        ws_sum.column_dimensions["B"].width = 20  # tested_at
        ws_sum.column_dimensions["C"].width = 16  # 반
        ws_sum.column_dimensions["D"].width = 14  # classId
        ws_sum.column_dimensions["E"].width = 10  # targetAge
        ws_sum.column_dimensions["F"].width = 22  # 대상명 (하이퍼링크)
        ws_sum.column_dimensions["G"].width = 14  # studentId
        ws_sum.column_dimensions["H"].width = 16  # loginIdUsed
        ws_sum.column_dimensions["I"].width = 16  # memNm
        ws_sum.column_dimensions["J"].width = 24  # memId
        ws_sum.column_dimensions["K"].width = 10  # 종류
        ws_sum.column_dimensions["L"].width = 10  # 총 API
        ws_sum.column_dimensions["M"].width = 10  # PASS
        ws_sum.column_dimensions["N"].width = 10  # FAIL
        ws_sum.column_dimensions["O"].width = 10  # 성공률
        ws_sum.freeze_panes = "A4"

        # ── 타깃별 시트 ──
        for meta in sorted_meta:
            label = meta["label"]
            results = results_by_target.get(label, [])
            sheet_name = sheet_name_map[label]
            ws = wb.create_sheet(title=sheet_name)

            t_pass = sum(1 for r in results if r[4] == "PASS")
            t_fail = sum(1 for r in results if r[4] == "FAIL")

            ws.merge_cells("A1:I1")
            sc = ws["A1"]
            sc.value = (
                f"{meta['type']}: {meta['mem_nm']} (memId={meta['mem_id']})  |  "
                f"서버: {server_label}  |  계정: {user_id}  |  "
                f"반: {meta.get('class_nm', '')} ({meta.get('class_id', '')})  |  "
                f"총 {len(results)}개  |  PASS: {t_pass}  |  FAIL: {t_fail}"
            )
            sc.font = Font(name="맑은 고딕", bold=True, size=12)
            sc.alignment = Alignment(vertical="center")
            ws.row_dimensions[1].height = 30

            col_headers = ["No.", "서버", "카테고리", "Method", "API Path", "Status", "결과", "Result Data", "에러 상세"]
            ws.append([])
            for ci, h in enumerate(col_headers, 1):
                c = ws.cell(row=3, column=ci, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = header_align
                c.border = thin_border
            ws.row_dimensions[3].height = 25

            for i, (tag, method, path, status, mark, rd, ed) in enumerate(results, 1):
                row = i + 3
                vals = [i, server_label, tag, method, path, status if status else "", mark, rd, ed]
                for ci, v in enumerate(vals, 1):
                    c = ws.cell(row=row, column=ci, value=v)
                    c.font = cell_font
                    c.border = thin_border
                    c.alignment = center_align if ci in (1, 2, 4, 6, 7) else cell_align

                rc = ws.cell(row=row, column=7)
                if mark == "PASS":
                    rc.fill = pass_fill
                    rc.font = pass_font
                else:
                    rc.fill = fail_fill
                    rc.font = fail_font

            tag_colors = {}
            color_toggle = ["F2F7FB", "FFFFFF"]
            cidx = 0
            prev_tag = None
            for i, (tag, *_) in enumerate(results):
                if tag != prev_tag:
                    if tag not in tag_colors:
                        tag_colors[tag] = color_toggle[cidx % 2]
                        cidx += 1
                    prev_tag = tag
                row = i + 4
                bg = PatternFill(start_color=tag_colors[tag], end_color=tag_colors[tag], fill_type="solid")
                for col in range(1, 8):
                    c = ws.cell(row=row, column=col)
                    if col != 7:
                        c.fill = bg

            ws.column_dimensions["A"].width = 6
            ws.column_dimensions["B"].width = 9
            ws.column_dimensions["C"].width = 16
            ws.column_dimensions["D"].width = 9
            ws.column_dimensions["E"].width = 48
            ws.column_dimensions["F"].width = 9
            ws.column_dimensions["G"].width = 9
            ws.column_dimensions["H"].width = 60
            ws.column_dimensions["I"].width = 60

            if results:
                ws.auto_filter.ref = f"A3:I{len(results) + 3}"
            ws.freeze_panes = "A4"

        wb.save(file_path)
        print(f"\n[INFO] 엑셀 저장 완료: {file_path}")
        print(f"[INFO] 요약 시트 1개 + 타깃 시트 {len(sorted_meta)}개")

    except Exception as e:
        print(f"[ERROR] ALL API test exception: {e!r}")
        print(traceback.format_exc())


def worker_api_pipeline(log_queue, user_id, user_pwd, server, device_label, steps,
                        gui_ctx=None):
    """ALL 또는 개별 API 선택 모두 worker_all_api_test로 라우팅하여 통일된 보고서 양식 생성."""
    import builtins

    def _print_via_queue(*args, sep=" ", end="\n", **kwargs):
        msg = sep.join(str(a) for a in args) + end
        log_queue.put(msg.rstrip("\n"))

    builtins.print = _print_via_queue

    # CALL_MAP과 동일 키 집합 + study/access (자동 수행)
    ALL_API_STEPS = [
        "수업시작 (study/access)",
        "커리큘럼 조회 (curriculum)",
        "출석 시간 전송 (attendance/curriculum)",
        "위티스쿨 메인 (witti-school/main)",
        "보유 위티팡 조회 (witti-app/main)",
        "아람북월드 과목 (aram-bookworld/subject)",
        "도서관 메인 (e-book/main)",
        "위티TV 메인 (tv/main)",
        "선생님 활동현황 (report/teacherActivityReport)",
        "학습 리포트 > 부모(학생) / 주간 (report/parentReport)",
    ]

    if "ALL" in steps:
        steps_to_run = ALL_API_STEPS
        print(f"[INFO] ALL 모드: 전체 API 실행 ({len(steps_to_run)}개)")
    else:
        # 개별 선택도 동일 양식 사용. 수업시작은 통과시킴(사용자가 명시 선택 시만 결과 기록됨)
        steps_to_run = list(steps)
        print(f"[INFO] 개별 모드: {len(steps_to_run)}개 API 실행")

    worker_all_api_test(log_queue, user_id, user_pwd, server, device_label, steps_to_run,
                        gui_ctx=gui_ctx)


if getattr(sys, "frozen", False):
    import __main__
    __main__.worker_main = worker_main
    __main__.worker_all_api_test = worker_all_api_test
    __main__.worker_api_pipeline = worker_api_pipeline
    print("Registered workers on __main__")


class MainApp(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.log_queue: Queue = None
        self.worker_process: Process = None
        # 이 인스턴스가 잡고 있는 기기 잠금. 테스트 실행 구간에만 유지한다.
        self.device_lock = DeviceLock()
        # 그 잠금을 소유한 워커 프로세스. _drain_timer는 테스트가 끝나도 계속
        # 돌기 때문에, '지금 잠금을 쥔 워커'를 따로 들고 있지 않으면 다음 실행에서
        # 잠금을 얻자마자 직전 워커가 죽어 있는 것을 보고 곧바로 풀어버린다.
        self._lock_proc = None
        self._drain_timer: QtCore.QTimer = None

        class EmittingStream(QtCore.QObject):
            textWritten = QtCore.pyqtSignal(str)
            def write(self, text):
                if not text or text == "\n": return
                self.textWritten.emit(text)
            def flush(self): pass
        self.stdout_stream = EmittingStream()
        sys.stdout = self.stdout_stream
        sys.stderr = self.stdout_stream
        self.stdout_stream.textWritten.connect(self.append_log)

        logging.basicConfig(level=logging.DEBUG, format="%(asctime)s [%(levelname)s] %(message)s", handlers=[logging.StreamHandler(sys.stdout)])
        self.logger = logging.getLogger(__name__)

        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setWindowFlag(QtCore.Qt.WindowStaysOnTopHint, True)
        ensure_adb_server()
        self.logger.info(describe_adb_server())
        self.load_devices()

        self.ui.listView_2.setGeometry(QtCore.QRect(10, 20, 291, 352))
        self.label_mem_id = QtWidgets.QLabel(self.ui.groupBox_11)
        self.label_mem_id.setGeometry(QtCore.QRect(10, 379, 291, 24))
        self.label_mem_id.setObjectName("label_mem_id")
        self.label_mem_id.setText("memId: -")
        self.label_auth_token = QtWidgets.QLabel(self.ui.groupBox_11)
        self.label_auth_token.setGeometry(QtCore.QRect(10, 407, 291, 24))
        self.label_auth_token.setObjectName("label_auth_token")
        self.label_auth_token.setText("authToken: -")

        self.class_list_data = []
        self.class_auth_token = None
        self.class_api_server = None
        self.study_access_mem_nm = None
        self.study_access_mem_id = None
        self.study_access_auth_token = None
        self.selected_class_nm = None
        self.selected_class_id = None
        self.selected_target_age = None
        self.selected_student_nm = None
        self.selected_student_id = None
        self.class_list_model = QtGui.QStandardItemModel(self)
        self.student_list_model = QtGui.QStandardItemModel(self)
        self.ui.listView.setModel(self.class_list_model)
        self.ui.listView.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.ui.listView_2.setModel(self.student_list_model)
        self.ui.listView_2.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)

        self.ui.comboBox.setItemData(1, 0, QtCore.Qt.UserRole)   # ALL
        self.ui.comboBox.setItemData(2, 1, QtCore.Qt.UserRole)   # 한글
        self.ui.comboBox.setItemData(3, 2, QtCore.Qt.UserRole)   # 수학
        self.ui.comboBox.setItemData(4, 3, QtCore.Qt.UserRole)   # 창의
        # STEP: index1=ALL(0), index2=STEP1, index3=STEP2
        self.ui.comboBox_2.setItemData(1, 0, QtCore.Qt.UserRole)   # ALL
        self.ui.comboBox_2.setItemData(2, 1, QtCore.Qt.UserRole)   # STEP 1
        self.ui.comboBox_2.setItemData(3, 2, QtCore.Qt.UserRole)   # STEP 2
        # 호: index1=ALL(0), index2~14 = 1~13호
        self.ui.comboBox_3.setItemData(1, 0, QtCore.Qt.UserRole)   # ALL
        for _i in range(1, 14):
            self.ui.comboBox_3.setItemData(_i + 1, _i, QtCore.Qt.UserRole)
        local_cfg = load_local_config()
        self.ui.lineEdit.setText(local_cfg.get("USER_ID", ""))
        self.ui.lineEdit_2.setText(local_cfg.get("USER_PWD", ""))

        self.ui.pushButton.clicked.connect(self.close)
        self.ui.pushButton_2.clicked.connect(self.open_report_folder)
        self.ui.pushButton_3.clicked.connect(self.on_start)
        self.ui.pushButton_4.clicked.connect(self.on_stop)
        self.ui.pushButton_5.clicked.connect(self.on_start)
        self.ui.pushButton_6.clicked.connect(self.load_devices)
        self.ui.pushButton_7.clicked.connect(self.on_start)
        self.ui.pushButton_8.clicked.connect(self.clear_log)
        self.ui.pushButton_10.clicked.connect(self.on_load_class_list)
        self.ui.listView.clicked.connect(self.on_class_item_clicked)
        self.ui.listView_2.clicked.connect(self.on_student_item_clicked)

        # ── Device mirroring (scrcpy) panel setup ──
        self._setup_mirror_ui()

        # ── API pipeline tab setup ──
        self._init_api_pipeline_tab()
        self.ui.pushButton_api_add.clicked.connect(self._api_pipeline_add)
        self.ui.pushButton_api_remove.clicked.connect(self._api_pipeline_remove)
        self.ui.listWidget_api_available.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.ui.listWidget_api_pipeline.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.ui.listWidget_api_available.doubleClicked.connect(self._api_pipeline_add)
        self.ui.listWidget_api_pipeline.doubleClicked.connect(self._api_pipeline_remove)
        self.ui.pushButton_api_run.clicked.connect(self.on_run_api_pipeline)

    # ── Device mirroring (scrcpy) ──────────────────────────────────────
    # 패널 내부에서 미러링 화면이 차지할 수 있는 최대 영역 (mirror_group 기준 좌표)
    # 가로형 태블릿(16:10)이 높이(671)를 꽉 채우도록 폭을 1078로 설정 (671*1.6=1073.6)
    MIRROR_AREA = QtCore.QRect(10, 22, 1078, 671)
    # 로그 영역 시작 y좌표 (centralwidget 기준).
    # 미러링 패널/탭 영역 하단(y=711, _setup_mirror_ui에서 맞춤) + 여백 6px.
    LOG_TOP = 717

    # Win32 상수 (scrcpy 창 임베드용)
    GWL_STYLE = -16
    WS_CHILD = 0x40000000
    WS_POPUP = 0x80000000
    WS_CAPTION = 0x00C00000
    WS_THICKFRAME = 0x00040000
    WS_VISIBLE = 0x10000000
    # 임베드된 scrcpy 창 클릭 감지 → 키보드 포커스 전환용
    WM_PARENTNOTIFY = 0x0210
    WM_LBUTTONDOWN = 0x0201

    def _setup_mirror_ui(self):
        # 창을 늘리고 로그 영역을 하단 전체 폭으로 이동
        # (로그 높이는 창 크기를 따라감 → resizeEvent에서 _layout_log_area 호출)
        self.resize(1838, 1162)
        self._layout_log_area()

        # 좌측 컨트롤/탭 영역과 미러링 패널의 좌우 자리 교체:
        # 원래 좌측(x<730)에 있던 centralwidget 직속 위젯들을 오른쪽으로 밀고
        # Device Screen 패널을 왼쪽에 배치한다 (하단 로그 영역은 전체 폭이라 무관)
        dx = 1117  # 좌측 여백(10) + 패널 폭(1098) + 패널-탭 간격(9)
        for w in self.ui.centralwidget.children():
            if isinstance(w, QtWidgets.QWidget) and w.x() < 730:
                w.move(w.x() + dx, w.y())

        # 탭 영역 하단(기존 y=791)이 미러링 패널 하단(y=711)과 일치하도록
        # 리스트 표시 영역(class/student/api)의 높이를 축소
        shrink = 80
        self.ui.tabWidget.resize(self.ui.tabWidget.width(),
                                 self.ui.tabWidget.height() - shrink)
        # 리스트 그룹박스와 리스트 위젯 높이 축소
        for w in (self.ui.groupBox_10, self.ui.groupBox_11, self.ui.groupBox_12,
                  self.ui.groupBox_13, self.ui.listView, self.ui.listView_2,
                  self.ui.listWidget_api_available, self.ui.listWidget_api_pipeline):
            w.resize(w.width(), w.height() - shrink)
        # 리스트 아래에 배치된 위젯들은 같은 폭만큼 위로 이동
        for w in (self.label_mem_id, self.label_auth_token,
                  self.ui.groupBox_4, self.ui.groupBox_8, self.ui.pushButton_api_run):
            w.move(w.x(), w.y() - shrink)

        # Device Screen 패널 생성 (미러링 영역 + 좌우 여백 10px)
        self.mirror_group = QtWidgets.QGroupBox(self.ui.centralwidget)
        self.mirror_group.setGeometry(QtCore.QRect(10, 8, 1098, 703))
        self.mirror_group.setTitle("Device Screen")
        self.mirror_group.setObjectName("groupBox_mirror")
        self.mirror_placeholder = QtWidgets.QLabel(self.mirror_group)
        self.mirror_placeholder.setGeometry(self.MIRROR_AREA)
        self.mirror_placeholder.setAlignment(QtCore.Qt.AlignCenter)
        self.mirror_placeholder.setText("디바이스를 선택한 후 [미러링] 버튼을 눌러주세요.")
        self.mirror_placeholder.setStyleSheet("color: #808080; background-color: #1e1e1e;")

        # Select Device 그룹: 새로고침 버튼을 위로 줄이고 아래에 미러링 버튼 추가
        self.ui.pushButton_6.setGeometry(QtCore.QRect(449, 12, 71, 23))
        self.pushButton_mirror = QtWidgets.QPushButton(self.ui.groupBox_5)
        self.pushButton_mirror.setGeometry(QtCore.QRect(449, 37, 71, 23))
        self.pushButton_mirror.setObjectName("pushButton_mirror")
        self.pushButton_mirror.setText("미러링")
        self.pushButton_mirror.clicked.connect(self.on_start_mirror)

        self.scrcpy_proc = None
        self._mirror_hwnd = None
        self._mirror_title = None
        self._mirror_aspect = None
        self._mirror_find_timer = None
        self._mirror_find_tries = 0
        self._mirror_visible_ticks = 0
        self._mirror_watch_timer = None
        self._scrcpy_log_file = None
        self._scrcpy_log_path = None

    def _read_scrcpy_log_tail(self, max_chars=600):
        """scrcpy 출력 로그 파일의 끝부분을 읽어 반환한다."""
        try:
            if self._scrcpy_log_file is not None:
                self._scrcpy_log_file.flush()
            if self._scrcpy_log_path and os.path.exists(self._scrcpy_log_path):
                with open(self._scrcpy_log_path, "rb") as f:
                    data = f.read()
                return data.decode("utf-8", errors="ignore").strip()[-max_chars:]
        except Exception:
            pass
        return ""

    def _layout_log_area(self):
        """로그 영역을 창 하단까지 채우도록 배치한다 (창 리사이즈 시에도 호출)."""
        if not hasattr(self, "ui") or not hasattr(self.ui, "menubar"):
            return
        central_h = self.height() - self.ui.menubar.height() - self.statusBar().height()
        log_h = max(120, central_h - self.LOG_TOP - 10)
        self.ui.groupBox_6.setGeometry(QtCore.QRect(10, self.LOG_TOP, 1818, log_h))
        self.ui.plainTextEdit.setGeometry(QtCore.QRect(15, 20, 1788, log_h - 30))

    def resizeEvent(self, event):
        self._layout_log_area()
        super().resizeEvent(event)

    def _query_device_aspect(self, serial):
        """adb wm size로 디바이스 해상도를 조회해 가로/세로 비율을 반환한다."""
        try:
            out = run_adb(["-s", serial, "shell", "wm", "size"],
                          text=True, encoding="utf-8", errors="ignore", timeout=5)
            m = re.search(r"Override size:\s*(\d+)x(\d+)", out) \
                or re.search(r"Physical size:\s*(\d+)x(\d+)", out)
            if m:
                w, h = int(m.group(1)), int(m.group(2))
                if w > 0 and h > 0:
                    self.logger.info(f"디바이스 해상도: {w}x{h}")
                    return w / h
        except Exception as e:
            self.logger.warning(f"디바이스 해상도 조회 실패: {e!r}")
        return None

    def _fit_mirror_rect(self, aspect):
        """MIRROR_AREA 안에 aspect 비율을 유지한 최대 크기 사각형(중앙 정렬)을 계산한다."""
        area = self.MIRROR_AREA
        w = area.width()
        h = int(w / aspect)
        if h > area.height():
            h = area.height()
            w = int(h * aspect)
        x = area.x() + (area.width() - w) // 2
        y = area.y() + (area.height() - h) // 2
        return QtCore.QRect(x, y, w, h)

    def on_start_mirror(self):
        serial = self.ui.comboBox_4.currentData()
        device_label = self.ui.comboBox_4.currentText()
        if not serial:
            self.logger.error("미러링할 디바이스를 선택해주세요.")
            return

        scrcpy_path = get_scrcpy_path()
        if not scrcpy_path:
            self.logger.error("scrcpy.exe를 찾을 수 없습니다. 프로젝트의 scrcpy 폴더를 확인해주세요.")
            return

        self.stop_mirror()
        self.mirror_placeholder.setText(f"{device_label}\n미러링 연결 중...")

        # scrcpy도 adb 클라이언트를 통해 붙는다. 서버가 없으면 scrcpy가 직접
        # 띄우게 되고, 그 기동 시간만큼 미러링 시작이 늦어진다.
        ensure_adb_server()

        self._mirror_aspect = self._query_device_aspect(serial)
        self._mirror_title = f"WittiMirror_{os.getpid()}_{abs(hash(serial)) % 100000}"

        # 번들 adb를 쓰도록 지정해 scrcpy 내장 adb와의 서버 충돌 방지
        env = os.environ.copy()
        env["ADB"] = get_adb_path()
        # scrcpy가 띄우는 adb 클라이언트도 이 인스턴스의 서버를 보게 한다.
        # 빠지면 scrcpy만 공용 5037로 붙어 기기를 못 찾는다.
        env["ANDROID_ADB_SERVER_PORT"] = str(get_adb_port())
        # 창 활성화용 첫 클릭도 터치로 전달 (마우스 입력 우선).
        # scrcpy가 자체적으로도 켜지만 env로 한 번 더 강제한다.
        # (env가 우선권을 가져서 scrcpy 로그에 "Could not enable mouse focus
        #  clickthrough" WARN이 뜨는데, 기능은 env 값으로 켜져 있으므로 무해)
        env["SDL_MOUSE_FOCUS_CLICKTHROUGH"] = "1"

        # Wi-Fi/USB 동일 프로파일 사용.
        # (사내망 실측: 미러링 중 RTT 평균 9ms, 링크 433Mbps → 대역폭 제한 불필요.
        #  30fps 제한을 걸면 오히려 움직임이 끊겨 보임)
        # max-size=1280: 패널 표시 폭(1078px)과 1:1에 가깝게 맞춰 선명도 최대화.
        # 그 이상은 표시 시 축소되어 화질 이득 없이 인코딩 부하만 증가.
        cmd = [
            scrcpy_path,
            "-s", str(serial),
            f"--window-title={self._mirror_title}",
            "--window-borderless",
            "--window-x=-10000",
            "--window-y=-10000",
            "--max-size=1280",
            "--video-bit-rate=12M",
            "--no-audio",
            "--stay-awake",
            # 숫자/특수문자를 텍스트 주입 대신 키코드로 전송.
            # SetParent로 임베드된 자식 창은 SDL3가 포커스를 인식하지 못해
            # text input 이벤트가 발생하지 않으므로 (한글/영문만 입력되고
            # 숫자/특수문자 입력 불가) raw key event 경로로 우회한다.
            "--raw-key-events",
        ]
        # 출력은 임시 파일로 리다이렉트 (PIPE는 버퍼가 차면 scrcpy가 블로킹될 수 있음)
        self._scrcpy_log_path = os.path.join(
            tempfile.gettempdir(), f"scrcpy_mirror_{os.getpid()}.log")
        try:
            self._scrcpy_log_file = open(self._scrcpy_log_path, "wb")
            self.scrcpy_proc = subprocess.Popen(
                cmd,
                env=env,
                cwd=os.path.dirname(scrcpy_path),
                stdout=self._scrcpy_log_file,
                stderr=subprocess.STDOUT,
                creationflags=subprocess.CREATE_NO_WINDOW,
            )
        except Exception as e:
            self.logger.error(f"scrcpy 실행 실패: {e!r}")
            self.mirror_placeholder.setText("미러링 실행 실패")
            return

        self.logger.info(f"scrcpy 시작 (PID={self.scrcpy_proc.pid}, device={device_label})")
        self.mirror_group.setTitle(f"Device Screen - {device_label}")

        # scrcpy 창이 생성되면 패널 안으로 임베드 (200ms 간격, 최대 20초 대기)
        self._mirror_find_tries = 0
        self._mirror_visible_ticks = 0
        if self._mirror_find_timer is None:
            self._mirror_find_timer = QtCore.QTimer(self)
            self._mirror_find_timer.timeout.connect(self._try_embed_mirror)
        self._mirror_find_timer.start(200)

    def _try_embed_mirror(self):
        # 창이 뜨기 전에 프로세스가 죽었으면 에러 출력
        if self.scrcpy_proc is None or self.scrcpy_proc.poll() is not None:
            self._mirror_find_timer.stop()
            tail = self._read_scrcpy_log_tail()
            self.logger.error(f"scrcpy가 종료되었습니다. {tail}")
            self.stop_mirror()
            self.mirror_placeholder.setText("미러링 연결 실패 (로그 확인)")
            return

        user32 = ctypes.windll.user32
        hwnd = user32.FindWindowW(None, self._mirror_title)
        if not hwnd:
            self._mirror_find_tries += 1
            if self._mirror_find_tries > 100:  # 20초 초과
                self._mirror_find_timer.stop()
                self.logger.error("scrcpy 창을 찾지 못했습니다. 미러링을 중단합니다.")
                self.stop_mirror()
            return

        # scrcpy 창은 첫 프레임 수신 시점에 표시됨. 렌더러 초기화가 끝나기 전에
        # reparent/resize하면 scrcpy가 크래시(divide by zero)하므로,
        # 창이 보이고 나서 3틱(약 0.6초) 더 기다린 뒤 임베드한다.
        if not user32.IsWindowVisible(wintypes.HWND(hwnd)):
            self._mirror_find_tries += 1
            if self._mirror_find_tries > 100:
                self._mirror_find_timer.stop()
                self.logger.error("scrcpy 창이 표시되지 않았습니다. 미러링을 중단합니다.")
                self.stop_mirror()
            return
        self._mirror_visible_ticks += 1
        if self._mirror_visible_ticks < 3:
            return

        self._mirror_find_timer.stop()

        # scrcpy 창 크기(기기 화면 비율 반영)로 표시 비율 결정, 실패 시 wm size 값 사용
        rect = wintypes.RECT()
        user32.GetWindowRect(wintypes.HWND(hwnd), ctypes.byref(rect))
        win_w, win_h = rect.right - rect.left, rect.bottom - rect.top
        aspect = self._mirror_aspect or (16 / 10)
        if win_w > 50 and win_h > 50 and 0.2 < (win_w / win_h) < 5:
            aspect = win_w / win_h

        # Win32 SetParent로 scrcpy 창을 메인 윈도우 안(패널 위치)에 임베드
        # 주의: 부모는 반드시 최상위 창(self.winId())이어야 함.
        #  - Qt createWindowContainer 또는 자식 위젯(winId 강제 생성)에 붙이면
        #    Qt가 네이티브 창을 재구성하면서 scrcpy(SDL)가 종료됨
        try:
            style = user32.GetWindowLongW(wintypes.HWND(hwnd), self.GWL_STYLE)
            style = (style & ~(self.WS_POPUP | self.WS_CAPTION | self.WS_THICKFRAME)) \
                | self.WS_CHILD | self.WS_VISIBLE
            user32.SetWindowLongW(wintypes.HWND(hwnd), self.GWL_STYLE,
                                  ctypes.c_long(style & 0xFFFFFFFF))
            user32.SetParent(wintypes.HWND(hwnd), wintypes.HWND(int(self.winId())))
            geo = self._fit_mirror_rect(aspect)
            # mirror_group 내부 좌표 → 메인 윈도우 좌표로 변환
            top_left = self.mirror_group.mapTo(self, geo.topLeft())
            user32.MoveWindow(wintypes.HWND(hwnd), top_left.x(), top_left.y(),
                              geo.width(), geo.height(), True)
            self._mirror_hwnd = hwnd
            self._focus_mirror_window()  # 키보드 입력도 바로 전달되도록
            self.mirror_placeholder.hide()
            self.logger.info("미러링 화면이 패널에 연결되었습니다.")
        except Exception as e:
            self.logger.error(f"미러링 창 임베드 실패: {e!r}")
            self.stop_mirror()
            return

        # scrcpy 프로세스 종료(기기 분리 등) 감시
        if self._mirror_watch_timer is None:
            self._mirror_watch_timer = QtCore.QTimer(self)
            self._mirror_watch_timer.timeout.connect(self._watch_mirror_proc)
        self._mirror_watch_timer.start(1000)

    def _focus_mirror_window(self):
        """임베드된 scrcpy 창으로 키보드 포커스를 넘긴다.

        scrcpy 창은 다른 프로세스 소유라 SetFocus가 그냥은 실패하므로
        (SetFocus는 호출 스레드의 입력 큐에 붙은 창에만 동작),
        AttachThreadInput으로 두 스레드의 입력 큐를 잠시 연결한 뒤 옮긴다.
        """
        user32 = ctypes.windll.user32
        hwnd = self._mirror_hwnd
        if not hwnd or not user32.IsWindow(wintypes.HWND(hwnd)):
            return
        target_tid = user32.GetWindowThreadProcessId(wintypes.HWND(hwnd), None)
        my_tid = ctypes.windll.kernel32.GetCurrentThreadId()
        if target_tid == my_tid:
            user32.SetFocus(wintypes.HWND(hwnd))
            return
        if user32.AttachThreadInput(my_tid, target_tid, True):
            try:
                user32.SetFocus(wintypes.HWND(hwnd))
            finally:
                user32.AttachThreadInput(my_tid, target_tid, False)

    def nativeEvent(self, eventType, message):
        # 임베드된 scrcpy 창(자식)이 클릭되면 부모인 이 창에 WM_PARENTNOTIFY가
        # 오므로, 이때 키보드 포커스를 scrcpy로 넘긴다.
        # 다른 Qt 위젯을 클릭하면 Qt가 포커스를 되찾아가므로 자연스럽게 전환된다.
        if eventType == b"windows_generic_MSG" and getattr(self, "_mirror_hwnd", None):
            msg = wintypes.MSG.from_address(int(message))
            if msg.message == self.WM_PARENTNOTIFY \
                    and (msg.wParam & 0xFFFF) == self.WM_LBUTTONDOWN:
                self._focus_mirror_window()
        return super().nativeEvent(eventType, message)

    def _watch_mirror_proc(self):
        if self.scrcpy_proc is not None and self.scrcpy_proc.poll() is not None:
            tail = self._read_scrcpy_log_tail()
            self.logger.info(f"미러링이 종료되었습니다 (exit={self.scrcpy_proc.returncode}). {tail}")
            self.stop_mirror()

    def stop_mirror(self):
        if self._mirror_find_timer is not None:
            self._mirror_find_timer.stop()
        if self._mirror_watch_timer is not None:
            self._mirror_watch_timer.stop()
        self._mirror_hwnd = None
        if self.scrcpy_proc is not None:
            if self.scrcpy_proc.poll() is None:
                try:
                    self.scrcpy_proc.terminate()
                    self.scrcpy_proc.wait(timeout=3)
                except Exception:
                    try:
                        self.scrcpy_proc.kill()
                    except Exception:
                        pass
            self.scrcpy_proc = None
        if self._scrcpy_log_file is not None:
            try:
                self._scrcpy_log_file.close()
            except Exception:
                pass
            self._scrcpy_log_file = None
        if self._scrcpy_log_path and os.path.exists(self._scrcpy_log_path):
            try:
                os.remove(self._scrcpy_log_path)
            except Exception:
                pass
            self._scrcpy_log_path = None
        self.mirror_group.setTitle("Device Screen")
        self.mirror_placeholder.setText("디바이스를 선택한 후 [미러링] 버튼을 눌러주세요.")
        self.mirror_placeholder.show()

    def closeEvent(self, event):
        self.stop_mirror()
        self._release_device_lock()
        super().closeEvent(event)

    # ── Device mirroring 끝 ────────────────────────────────────────────

    def open_report_folder(self):
        project_dir = os.path.abspath(os.getcwd())
        report_dir  = os.path.join(project_dir, "test_report")
        if not os.path.isdir(report_dir):
            try:
                os.makedirs(report_dir, exist_ok=True)
            except Exception as e:
                self.logger.error(f"폴더 생성 실패: {e!r}")
                return
        QDesktopServices.openUrl(QUrl.fromLocalFile(report_dir))

    @QtCore.pyqtSlot()
    def clear_log(self):
        self.ui.plainTextEdit.clear()

    @QtCore.pyqtSlot(str)
    def append_log(self, text):
        edit = self.ui.plainTextEdit
        edit.appendPlainText(text.rstrip("\n"))
        edit.verticalScrollBar().setValue(edit.verticalScrollBar().maximum())

    @staticmethod
    def _resolve_api_server(server_name):
        mapping = {
            "Prod": "api",
            "QA": "qa-api",
            "Dev": "dev-api",
        }
        return mapping.get(server_name, "api")

    def on_load_class_list(self):
        input_id = self.ui.lineEdit.text().strip()
        input_pwd = self.ui.lineEdit_2.text().strip()
        server_name = self.ui.comboBox_6.currentText()
        api_server = self._resolve_api_server(server_name)

        if not input_id or not input_pwd:
            self.logger.error("Class List 조회를 위해 ID/PW를 입력해주세요.")
            return

        auth_token = login_step1(input_id, input_pwd, api_server)
        if not auth_token:
            self.logger.error("Class List 조회 실패: 로그인(authToken 발급) 실패")
            return
        self.class_auth_token = auth_token
        self.class_api_server = api_server

        response = class_list(auth_token, input_id, api_server)
        if response is None:
            self.logger.error("Class List 조회 실패: API 응답 없음")
            return

        try:
            data = response.json()
            classes = data.get("result", {}).get("classList", [])
            self.class_list_data = classes

            self.class_list_model.clear()
            self.student_list_model.clear()

            all_item = QtGui.QStandardItem("ALL")
            all_item.setData("", QtCore.Qt.UserRole)
            self.class_list_model.appendRow(all_item)

            for cls in classes:
                class_nm = str(cls.get("classNm", "")).strip() or "-"
                target_age = str(cls.get("targetAge", "")).strip() or "-"
                class_id = str(cls.get("classId", "")).strip()
                item = QtGui.QStandardItem(f"{class_nm} / {target_age}")
                item.setData(class_id, QtCore.Qt.UserRole)
                self.class_list_model.appendRow(item)

            self.logger.info(f"Class List {len(classes)}건 로드 완료")
            self._auto_select_first_class_and_student()
        except Exception as e:
            self.logger.error(f"Class List 파싱 실패: {e!r}")

    def _auto_select_first_class_and_student(self):
        if self.class_list_model.rowCount() <= 1:
            self.logger.warning("Auto select skipped: class list is empty.")
            return

        # 기본 선택은 최상위 "ALL" (index 0) — 전체 반 순회가 기본 동작
        all_index = self.class_list_model.index(0, 0)
        if not all_index.isValid():
            self.logger.warning("Auto select skipped: invalid ALL index.")
            return

        self.ui.listView.setCurrentIndex(all_index)
        self.on_class_item_clicked(all_index)

        if self.student_list_model.rowCount() <= 0:
            self.logger.warning("Auto select skipped: student list is empty.")
            return

        first_student_index = self.student_list_model.index(0, 0)
        if not first_student_index.isValid():
            self.logger.warning("Auto select skipped: invalid first student index.")
            return

        self.ui.listView_2.setCurrentIndex(first_student_index)
        self.on_student_item_clicked(first_student_index)
        self.logger.info("Auto-selected ALL class and first student.")

    def on_class_item_clicked(self, index):
        class_id = index.data(QtCore.Qt.UserRole)
        if class_id is None:
            self.logger.warning("선택한 클래스의 classId를 찾을 수 없습니다.")
            return

        if class_id == "":
            # "ALL" 선택: 특정 반이 아닌 전체 반 순회를 의미.
            # 단, study/access 토큰 확보용으로 첫 반의 학생 목록은 그대로 채워둔다.
            self.selected_class_id = ""
            self.selected_class_nm = "ALL"
            self.selected_target_age = ""
            self.logger.info("클래스 선택: ALL (전체 반 대상)")
            first_class_id = ""
            if self.class_list_model.rowCount() > 1:
                first_class_id = str(
                    self.class_list_model.index(1, 0).data(QtCore.Qt.UserRole) or ""
                ).strip()
            if first_class_id:
                self._load_student_list(first_class_id)
            else:
                self.student_list_model.clear()
            return

        # 선택된 class 정보 저장
        self.selected_class_id = class_id
        display = index.data(QtCore.Qt.DisplayRole) or ""
        self.selected_class_nm = display.split(" / ")[0] if " / " in display else display
        self.selected_target_age = display.split(" / ")[1] if " / " in display else ""
        self._load_student_list(class_id)

    def _load_student_list(self, class_id):
        if not self.class_auth_token or not self.class_api_server:
            self.logger.error("학생 목록 조회 실패: authToken/server 정보가 없습니다. 먼저 Class List를 조회해주세요.")
            return

        response = student_list_by_class(self.class_auth_token, class_id, self.class_api_server)
        if response is None:
            self.logger.error("학생 목록 조회 실패: API 응답 없음")
            return

        try:
            data = response.json()
            result = data.get("result", {})
            students = result.get("studentList", [])
            teacher_mem_id = str(result.get("teacherMemId", "")).strip()
            teacher_mem_nm = str(
                result.get("teacherMemNm")
                or result.get("teacherNm")
                or result.get("memNm")
                or ""
            ).strip()
            students = sorted(
                students,
                key=lambda s: str(s.get("studentNm", "")).strip(),
            )

            self.student_list_model.clear()

            # teacherMemId가 있으면 맨 위에 "선생님(memNm)" 항목 추가
            if teacher_mem_id:
                teacher_label = f"선생님({teacher_mem_nm})" if teacher_mem_nm else "선생님"
                teacher_login_id = self.ui.lineEdit.text().strip()
                teacher_data = {
                    "studentNm": teacher_label,
                    "studentId": teacher_mem_id,
                    "loginId": teacher_login_id,
                    "isTeacher": True,
                }
                t_item = QtGui.QStandardItem(teacher_label)
                t_item.setData(teacher_mem_id, QtCore.Qt.UserRole)
                t_item.setData(teacher_data, QtCore.Qt.UserRole + 1)
                self.student_list_model.appendRow(t_item)

            for student in students:
                student_nm = str(student.get("studentNm", "")).strip() or "-"
                student_id = str(student.get("studentId", "")).strip()
                item = QtGui.QStandardItem(student_nm)
                item.setData(student_id, QtCore.Qt.UserRole)
                item.setData(student, QtCore.Qt.UserRole + 1)
                self.student_list_model.appendRow(item)

            self.logger.info(
                f"Student List {len(students)}건 로드 완료 "
                f"(classId={class_id}, teacherMemId={teacher_mem_id or '없음'})"
            )
        except Exception as e:
            self.logger.error(f"Student List 파싱 실패: {e!r}")

    def on_student_item_clicked(self, index):
        student_id = str(index.data(QtCore.Qt.UserRole) or "").strip()
        student_data = index.data(QtCore.Qt.UserRole + 1) or {}
        login_id = str(
            student_data.get("loginId")
            or student_data.get("studentLoginId")
            or self.ui.lineEdit.text().strip()
        ).strip()

        if not student_id:
            self.logger.warning("선택한 학생의 studentId를 찾을 수 없습니다.")
            return
        # 선택된 student 정보 저장
        self.selected_student_id = student_id
        self.selected_student_nm = str(student_data.get("studentNm", "")).strip()
        if not login_id:
            self.logger.error("study/access 호출 실패: loginId를 찾을 수 없습니다.")
            return

        server_name = self.ui.comboBox_6.currentText()
        api_server = self._resolve_api_server(server_name)
        response = authenticate_study_access(student_id, login_id, api_server)
        if response is None:
            self.logger.error(
                f"study/access 호출 실패: studentId={student_id}, loginId={login_id}, "
                f"studentKeys={list(student_data.keys()) if isinstance(student_data, dict) else 'N/A'}"
            )
            return

        mem_nm, mem_id, auth_token = get_study_access_auth()
        self.study_access_mem_nm = mem_nm
        self.study_access_mem_id = mem_id
        self.study_access_auth_token = auth_token
        token_masked = f"{auth_token[:12]}..." if auth_token else "None"
        self.label_mem_id.setText(f"memId: {mem_id if mem_id else '-'}")
        self.label_auth_token.setText(f"authToken: {token_masked}")
        self.logger.info(
            f"study/access 완료 (studentId={student_id}, memId={mem_id}, authToken={token_masked})"
        )

    def load_devices(self):
        # 서버가 없으면 아래 run_adb의 adb 클라이언트가 자기가 띄우는데,
        # 그 기동 비용(실측 약 2초)을 이 함수가 통째로 뒤집어쓴다.
        # 여기서 먼저 확보한다. (이미 떠 있으면 아무 동작도 안 한다)
        ensure_adb_server()
        try:
            out = run_adb(["devices", "-l"], text=True, encoding="utf-8", errors="ignore", timeout=20)
            lines = [ln for ln in out.strip().splitlines()[1:] if ln.strip()]
            entries = []
            for ln in lines:
                parts = ln.split()
                if len(parts) < 2 or parts[1] != "device": continue
                dev_id = parts[0]
                model = next((p.split(":", 1)[1] for p in parts if p.startswith("model:")), "").replace("_", "-")
                wifi = False
                serial_hint = ""
                m = re.search(r"^adb-([A-Za-z0-9]+)-", dev_id)
                if m and "._adb-tls-connect._tcp" in dev_id:
                    serial_hint = m.group(1)
                    wifi = True
                elif ":" in dev_id and dev_id.rsplit(":", 1)[1].isdigit():
                    wifi = True
                if serial_hint: canon = serial_hint
                else:
                    try: canon = run_adb(["-s", dev_id, "get-serialno"], text=True, timeout=3).strip()
                    except Exception: canon = dev_id
                entries.append({"dev_id": dev_id, "model": model, "wifi": wifi, "canon": canon})

            by_serial = {}
            for e in entries:
                k = e["canon"]
                by_serial.setdefault(k, {})
                by_serial[k]['wifi' if e['wifi'] else 'usb'] = e

            items = []
            for k, d in by_serial.items():
                e = d.get('usb') or d.get('wifi')
                model, aos = e['model'], ""
                try:
                    props = run_adb(["-s", e['dev_id'], "shell",
                                     "getprop ro.product.model; getprop ro.build.version.release"],
                                    text=True, timeout=3).strip().splitlines()
                    if props: model = props[0].strip() or model
                    if len(props) >= 2: aos = props[1].strip()
                except Exception:
                    pass
                display_name = model or k
                if aos: display_name += f" / AOS {aos}"
                display_name += f" ({k})"
                if 'usb' in d: items.append((f"{display_name} [USB]", d['usb']['canon'], k))
                if 'wifi' in d: items.append((f"{display_name} [Wi-Fi]", d['wifi']['dev_id'], k))
            if not items: items = [("(no devices)", "", "")]
        except Exception as e:
            items = [(f"Error: {e}", "", "")]

        self.ui.comboBox_4.clear()
        for label, dev_id, canon in items:
            # 다른 인스턴스가 테스트 중인 기기는 라벨에 표시한다.
            # USB/Wi-Fi 어느 경로로 보이든 같은 기기면 같은 잠금에 걸린다.
            if canon and canon != self.device_lock.serial:
                label += describe_holder(canon)
            self.ui.comboBox_4.addItem(label, dev_id)
            self.ui.comboBox_4.setItemData(
                self.ui.comboBox_4.count() - 1, canon, QtCore.Qt.UserRole + 1)

    def on_start(self):
        if self.worker_process and self.worker_process.is_alive():
            self.logger.warning("작업이 이미 실행 중입니다.")
            return

        device_name = self.ui.comboBox_4.currentData()
        device_label = self.ui.comboBox_4.currentText()
        inputId     = self.ui.lineEdit.text().strip()
        inputPwd    = self.ui.lineEdit_2.text().strip()
        subjCd      = self.ui.comboBox.currentData()
        itemCd      = self.ui.comboBox_2.currentData()
        curtnSeq    = self.ui.comboBox_3.currentData()
        btn_name    = self.sender().objectName()
        title_name  = self.ui.comboBox_5.currentText()
        server      = self.ui.comboBox_6.currentText()

        if not inputId or not inputPwd:
            self.logger.error("ID와 PWD를 모두 입력해주세요.")
            return
        if btn_name == "pushButton_3" and (self.ui.comboBox.currentIndex()==0 or self.ui.comboBox_2.currentIndex()==0 or self.ui.comboBox_3.currentIndex()==0):
            self.logger.error("과목, STEP, 호를 모두 선택해주세요.")
            return
        if btn_name == "pushButton_7" and self.ui.comboBox_5.currentIndex()==0:
            self.logger.error("Song을 선택해주세요.")
            return
        if btn_name in {"pushButton_5", "pushButton_3", "pushButton_7"} and not self.study_access_auth_token:
            self.logger.error("study/access authToken이 없습니다. 먼저 학생을 선택해주세요.")
            return

        # 같은 기기를 다른 인스턴스가 이미 테스트 중이면 시작하지 않는다.
        # 무선 연결은 여러 인스턴스가 동시에 붙을 수 있어 adb가 막아주지 못한다.
        if not self._acquire_device_lock():
            return

        self.log_queue = Queue()
        args = (
            self.log_queue,
            btn_name,
            device_name,
            device_label,
            inputId,
            inputPwd,
            subjCd,
            itemCd,
            curtnSeq,
            title_name,
            server,
            self.study_access_mem_nm,
            self.study_access_mem_id,
            self.study_access_auth_token,
            str(self.selected_class_id or "").strip(),
        )
        self.worker_process = Process(target=worker_main, args=args)
        try:
            self.worker_process.start()
        except Exception:
            # 프로세스가 뜨지 못하면 잠금이 영원히 남는다
            self._release_device_lock()
            raise
        self._lock_proc = self.worker_process
        self.logger.info(f"AutoTest 프로세스 시작 (PID={self.worker_process.pid})")

        if self._drain_timer is None:
            self._drain_timer = QtCore.QTimer(self)
            self._drain_timer.timeout.connect(self._drain_logs)
        self._drain_timer.start(100)

    # ── API pipeline helpers ───────────────────────────────────────────
    API_STEPS = [
        "ALL",
        "수업시작 (study/access)",
        "커리큘럼 조회 (curriculum)",
        "출석 시간 전송 (attendance/curriculum)",
        "위티스쿨 메인 (witti-school/main)",
        "보유 위티팡 조회 (witti-app/main)",
        "아람북월드 과목 (aram-bookworld/subject)",
        "도서관 메인 (e-book/main)",
        "위티TV 메인 (tv/main)",
        "선생님 활동현황 (report/teacherActivityReport)",
        "학습 리포트 > 부모(학생) / 주간 (report/parentReport)",
    ]

    def _init_api_pipeline_tab(self):
        """좌측 API 목록 리스트에 사용 가능한 API 항목을 채운다."""
        for name in self.API_STEPS:
            self.ui.listWidget_api_available.addItem(name)

    def _api_pipeline_add(self):
        """좌측에서 선택한 API를 우측 리스트에 추가 (중복 불가)."""
        item = self.ui.listWidget_api_available.currentItem()
        if item is None:
            return
        name = item.text()

        # ALL이 이미 있으면 다른 항목 추가 불가
        for i in range(self.ui.listWidget_api_pipeline.count()):
            existing = self.ui.listWidget_api_pipeline.item(i).text().split(". ", 1)[-1]
            if existing == "ALL" and name != "ALL":
                return
        # ALL을 추가하면 기존 항목 모두 비우고 ALL만 남김
        if name == "ALL":
            self.ui.listWidget_api_pipeline.clear()
            self.ui.listWidget_api_pipeline.addItem("1. ALL")
            return

        # 중복 체크
        for i in range(self.ui.listWidget_api_pipeline.count()):
            existing = self.ui.listWidget_api_pipeline.item(i).text()
            if existing.split(". ", 1)[-1] == name:
                return
        idx = self.ui.listWidget_api_pipeline.count() + 1
        self.ui.listWidget_api_pipeline.addItem(f"{idx}. {name}")

    def _api_pipeline_remove(self):
        """우측 파이프라인 리스트에서 선택 항목을 제거하고 번호 재정렬."""
        row = self.ui.listWidget_api_pipeline.currentRow()
        if row < 0:
            return
        self.ui.listWidget_api_pipeline.takeItem(row)
        # 번호 재정렬
        for i in range(self.ui.listWidget_api_pipeline.count()):
            text = self.ui.listWidget_api_pipeline.item(i).text()
            # "N. 실제이름" → "새번호. 실제이름"
            name = text.split(". ", 1)[-1]
            self.ui.listWidget_api_pipeline.item(i).setText(f"{i + 1}. {name}")

    def _get_pipeline_steps(self):
        """우측 리스트에서 API 이름만 순서대로 추출."""
        steps = []
        for i in range(self.ui.listWidget_api_pipeline.count()):
            text = self.ui.listWidget_api_pipeline.item(i).text()
            name = text.split(". ", 1)[-1]
            steps.append(name)
        return steps

    def on_run_api_pipeline(self):
        """파이프라인 실행 버튼 클릭 핸들러."""
        if self.worker_process and self.worker_process.is_alive():
            self.logger.warning("작업이 이미 실행 중입니다.")
            return

        steps = self._get_pipeline_steps()
        if not steps:
            msg_box = QMessageBox()
            msg_box.setIcon(QMessageBox.Warning)
            msg_box.setText("실행할 API를 파이프라인에 추가해주세요.")
            msg_box.setWindowTitle("파이프라인 비어있음")
            msg_box.exec_()
            return

        user_id = self.ui.lineEdit.text().strip()
        user_pwd = self.ui.lineEdit_2.text().strip()
        server = self.ui.comboBox_6.currentText()

        if not all([user_id, user_pwd, server]):
            msg_box = QMessageBox()
            msg_box.setIcon(QMessageBox.Warning)
            msg_box.setText("ID, PW, 서버를 모두 입력해주세요.")
            msg_box.setWindowTitle("입력 오류")
            msg_box.exec_()
            return

        self.log_queue = Queue()
        device_label = self.ui.comboBox_4.currentText()
        gui_ctx = {
            "auth_token": self.study_access_auth_token,
            "mem_id": self.study_access_mem_id,
            "class_nm": self.selected_class_nm,
            "class_id": self.selected_class_id,
            "target_age": self.selected_target_age,
            "student_nm": self.selected_student_nm,
            "student_id": self.selected_student_id,
        }
        args = (self.log_queue, user_id, user_pwd, server, device_label, steps, gui_ctx)
        self.worker_process = Process(target=worker_api_pipeline, args=args)
        self.worker_process.start()
        self.logger.info(f"API pipeline 프로세스 시작 (PID={self.worker_process.pid}), steps={steps}")

        if self._drain_timer is None:
            self._drain_timer = QtCore.QTimer(self)
            self._drain_timer.timeout.connect(self._drain_logs)
        self._drain_timer.start(100)

    def _drain_logs(self):
        if self.log_queue:
            while not self.log_queue.empty():
                try:
                    line = self.log_queue.get_nowait()
                    self.append_log(line)
                except Exception:
                    break

        # 잠금을 쥔 워커가 끝났으면 놓아준다.
        # 남은 로그를 모두 비운 뒤에 확인해야 마지막 줄이 잘리지 않는다.
        if self._lock_proc is not None and not self._lock_proc.is_alive():
            self._release_device_lock()

    def _acquire_device_lock(self):
        """선택한 기기의 잠금을 얻는다. 다른 인스턴스가 쓰는 중이면 False."""
        idx = self.ui.comboBox_4.currentIndex()
        # 잠금 키는 정식 시리얼(canon)이어야 한다. USB 항목과 Wi-Fi 항목은
        # dev_id가 서로 다르므로, dev_id로 잠그면 무선으로 우회 실행된다.
        canon = self.ui.comboBox_4.itemData(idx, QtCore.Qt.UserRole + 1) or             self.ui.comboBox_4.currentData()
        if not canon:
            self.logger.error("디바이스를 선택해주세요.")
            return False

        label = self.ui.comboBox_4.currentText()
        if self.device_lock.acquire(canon, instance=os.path.basename(_instance_dir()),
                                    device_label=label):
            return True

        self.logger.error(
            f"이 기기는 다른 인스턴스가 사용 중입니다{describe_holder(canon)}. "
            f"해당 테스트가 끝난 뒤 [새로고침] 후 다시 시도해주세요.")
        return False

    def _release_device_lock(self):
        self._lock_proc = None
        if self.device_lock.held:
            self.logger.info(f"기기 잠금 해제: {self.device_lock.serial}")
            self.device_lock.release()

    def on_stop(self):
        if self.worker_process and self.worker_process.is_alive():
            self.worker_process.terminate()
            self.worker_process.join(1)
            self.logger.info("작업 프로세스가 중단되었습니다.")
        else:
            self.logger.info("실행 중인 작업이 없습니다.")
        if self._drain_timer:
            self._drain_timer.stop()
        # 타이머를 멈추면 _drain_logs가 더 이상 돌지 않으므로 여기서 직접 놓아준다
        self._release_device_lock()


def main():
    app = QApplication(sys.argv)
    window = MainApp()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    freeze_support()
    main()
