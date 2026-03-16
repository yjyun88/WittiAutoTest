"""
전체 API 응답 테스트 스크립트 (인증/클래스 선택 제외)
테스트 계정: MGtest000 / mini1122@@
"""
import requests
import base64
import json
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

SERVER = "dev-api"
BASE = f"https://{SERVER}.wittiverse.com/v2"
USER_ID = "MGtest000"
USER_PWD = "mini1122@@"
X_DEVICE_INFO = "QW5kcm9pZC4zMzo6OlI5VFgyMDJHNUFFTTo6OlI5VFgyMDJHNUFFTTo6OktOT1g6OjpTTS1YMjE2Ojo6YXBwLjE0Ojo6"

results = []

def log(tag, method, path, r):
    status = r.status_code if r else 0
    ok = r and r.ok
    mark = "PASS" if ok else "FAIL"
    result_data = ""
    error_detail = ""
    if r is not None:
        try:
            body = r.text[:1000]
            if ok:
                result_data = body
            else:
                error_detail = body
        except:
            pass
    results.append((tag, method, path, status, mark, result_data, error_detail))
    print(f"  [{mark}] {method:6s} {path} -> {status} {error_detail[:120]}")

def safe_request(method, path, headers=None, params=None, json_body=None):
    url = f"{BASE}{path}"
    try:
        resp = requests.request(method, url, headers=headers, params=params, json=json_body, timeout=20)
        return resp
    except Exception as e:
        print(f"    [EXCEPTION] {method} {path}: {e}")
        return None

# ── Step 0: 로그인 ──
print("=== 로그인 ===")
enc_id = base64.b64encode(USER_ID.encode()).decode()
enc_pwd = base64.b64encode(USER_PWD.encode()).decode()
login_resp = safe_request("POST", "/authenticate/login",
    headers={"Content-Type": "application/json", "X-device-info": X_DEVICE_INFO},
    params={"loginTp": "L", "ptnrId": "1102"},
    json_body={"loginId": enc_id, "loginPwd": enc_pwd})

if not login_resp or not login_resp.ok:
    print(f"로그인 실패: {login_resp.status_code if login_resp else 'No response'}")
    if login_resp:
        print(login_resp.text[:300])
    exit(1)

teacher_token = login_resp.json().get("result", {}).get("authToken")
print(f"교사 토큰 획득: {teacher_token[:20]}...")

# 클래스/학생 조회하여 첫 번째 학생으로 study/access
class_resp = safe_request("GET", f"/authenticate/classes/{USER_ID}",
    headers={"Authorization": f"Bearer {teacher_token}"})
classes = class_resp.json().get("result", {}).get("classList", [])
first_class = classes[0] if classes else {}
class_id = str(first_class.get("classId", ""))
target_age = str(first_class.get("targetAge", ""))
print(f"첫 번째 클래스: {first_class.get('classNm')} (classId={class_id}, targetAge={target_age})")

stu_resp = safe_request("GET", "/authenticate/classes",
    headers={"Authorization": f"Bearer {teacher_token}"},
    params={"classId": class_id})
students = stu_resp.json().get("result", {}).get("studentList", [])
first_stu = students[0] if students else {}
student_id = str(first_stu.get("studentId", ""))
print(f"첫 번째 학생: {first_stu.get('studentNm')} (studentId={student_id})")

# study/access로 자녀 토큰 획득
access_resp = safe_request("POST", "/authenticate/study/access",
    headers={"Content-Type": "application/json", "X-device-info": X_DEVICE_INFO},
    params={"loginTp": "L", "autoLoginYn": "Y", "ptnrId": "1000"},
    json_body={"studentId": student_id, "loginId": USER_ID, "accessType": "C"})

access_data = access_resp.json().get("result", {})
child_token = access_data.get("authToken", "")
child_id = str(access_data.get("memId", ""))
print(f"자녀 토큰 획득: {child_token[:20]}...")
print(f"childId(memId): {child_id}")

# 커리큘럼에서 실제 prodId 추출
curriculum_resp = safe_request("GET", "/witti-box/curriculum",
    headers={"Authorization": f"Bearer {child_token}"},
    params={"childId": child_id})
real_prod_id = "P0001"
real_conts_id = None
if curriculum_resp and curriculum_resp.ok:
    cur_data = curriculum_resp.json().get("result", {})
    # wittiSchoolList에서 첫 번째 prodId 추출
    school_list = cur_data.get("wittiSchoolList", [])
    if school_list:
        real_prod_id = school_list[0].get("prodId", "P0001")
        print(f"실제 prodId: {real_prod_id}")
    # wittiTvList에서 첫 번째 contsId 추출
    tv_list = cur_data.get("wittiTvList", [])
    if tv_list:
        real_conts_id = tv_list[0].get("contsId")
        print(f"실제 TV contsId: {real_conts_id}")

# TV 카테고리에서 실제 catgryCd 추출
AUTH = {"Authorization": f"Bearer {child_token}"}
AUTH_JSON = {"Authorization": f"Bearer {child_token}", "Content-Type": "application/json"}
AUTH_DEVICE = {**AUTH, "X-device-info": X_DEVICE_INFO}
AUTH_JSON_DEVICE = {**AUTH_JSON, "X-device-info": X_DEVICE_INFO}
TEACHER_AUTH = {"Authorization": f"Bearer {teacher_token}"}
TEACHER_AUTH_JSON = {"Authorization": f"Bearer {teacher_token}", "Content-Type": "application/json"}

tv_main_resp = safe_request("GET", "/tv/main", headers=AUTH)
real_tv_catgry = None
real_tv_conts_id = None
if tv_main_resp and tv_main_resp.ok:
    tv_data = tv_main_resp.json().get("result", {})
    cat_list = tv_data.get("categoryList") or tv_data.get("catgryList") or []
    if cat_list:
        real_tv_catgry = cat_list[0].get("catgryCd")
        print(f"실제 TV catgryCd: {real_tv_catgry}")
    # contsId 추출 시도
    for key in ["newList", "popularList", "recentList", "contentsList"]:
        items = tv_data.get(key, [])
        if items:
            real_tv_conts_id = items[0].get("contsId")
            if real_tv_conts_id:
                print(f"실제 TV contsId: {real_tv_conts_id}")
                break

tv_cat = real_tv_catgry or "L3000"

# witti-school/main에서 실제 prodId/ptnrId 추출
school_main_resp = safe_request("GET", "/witti-school/main", headers=AUTH)
real_school_prod = real_prod_id
real_ptnr_id = "1102"
if school_main_resp and school_main_resp.ok:
    sm_data = school_main_resp.json().get("result", {})
    prod_list = sm_data.get("prodList") or sm_data.get("productList") or []
    if prod_list:
        real_school_prod = prod_list[0].get("prodId", real_prod_id)
        real_ptnr_id = str(prod_list[0].get("ptnrId", "1102"))
        print(f"실제 school prodId: {real_school_prod}, ptnrId: {real_ptnr_id}")

# 미니게임 contsId 추출 시도 (커리큘럼의 wittiPlayList)
real_game_conts = "G0001"
if curriculum_resp and curriculum_resp.ok:
    cur_data = curriculum_resp.json().get("result", {})
    play_list = cur_data.get("wittiPlayList", [])
    if play_list:
        real_game_conts = play_list[0].get("contsId", "G0001")
        print(f"실제 game contsId: {real_game_conts}")

print("\n" + "=" * 60)
print(f"테스트 시작 (서버: {SERVER})")
print("=" * 60)

# ── 위티스쿨 ──
print("\n=== 위티스쿨 ===")

r = safe_request("GET", "/witti-school/verify", headers=AUTH)
log("위티스쿨", "GET", "/witti-school/verify", r)

r = safe_request("GET", "/witti-school/brand", headers=AUTH)
log("위티스쿨", "GET", "/witti-school/brand", r)

r = safe_request("GET", "/witti-school/main", headers=AUTH)
log("위티스쿨", "GET", "/witti-school/main", r)

r = safe_request("GET", "/witti-school/aram-bookworld/main", headers=AUTH,
    params={"ptnrId": real_ptnr_id, "prodId": real_school_prod})
log("위티스쿨", "GET", "/witti-school/aram-bookworld/main", r)

r = safe_request("GET", "/witti-school/aram-bookworld/subject", headers=AUTH,
    params={"ptnrId": real_ptnr_id, "prodId": real_school_prod})
log("위티스쿨", "GET", "/witti-school/aram-bookworld/subject", r)

r = safe_request("GET", "/witti-school/aram-bookworld/subject/KOR", headers=AUTH,
    params={"ptnrId": real_ptnr_id, "prodId": real_school_prod})
log("위티스쿨", "GET", "/witti-school/aram-bookworld/subject/{subjCd}", r)

r = safe_request("GET", "/witti-school/e-book/main", headers=AUTH)
log("위티스쿨", "GET", "/witti-school/e-book/main", r)

r = safe_request("GET", "/witti-school/e-book/main/book/list/catgry", headers=AUTH,
    params={"catgryCd": "ALL", "pageNo": 1, "pageRow": 10})
log("위티스쿨", "GET", "/witti-school/e-book/main/book/list/catgry", r)

r = safe_request("POST", "/witti-school/start-study", headers=AUTH_JSON,
    json_body={"prodId": real_school_prod, "startDate": datetime.now().strftime("%Y-%m-%d")})
log("위티스쿨", "POST", "/witti-school/start-study", r)

r = safe_request("POST", "/witti-school/reset-study", headers=AUTH_JSON,
    json_body={"prodId": real_school_prod})
log("위티스쿨", "POST", "/witti-school/reset-study", r)

# ── 위티TV ──
print("\n=== 위티TV ===")

r = safe_request("GET", "/tv/verify", headers=AUTH)
log("위티TV", "GET", "/tv/verify", r)

r = safe_request("GET", "/tv/category", headers=AUTH, params={"catgryCd": tv_cat})
log("위티TV", "GET", "/tv/category", r)

r = safe_request("GET", "/tv/main", headers=AUTH)
log("위티TV", "GET", "/tv/main", r)

r = safe_request("GET", f"/tv/main/{tv_cat}", headers=AUTH)
log("위티TV", "GET", "/tv/main/{catgryCd}", r)

r = safe_request("GET", "/tv/category/detail", headers=AUTH, params={"catgryCd": tv_cat})
log("위티TV", "GET", "/tv/category/detail", r)

r = safe_request("GET", "/tv/category/detail/newPop", headers=AUTH,
    params={"topCatgryCd": tv_cat, "upCatgryCd": tv_cat, "newPopTp": "new", "searchTp": "new"})
log("위티TV", "GET", "/tv/category/detail/newPop", r)

r = safe_request("GET", "/tv/favorites", headers=AUTH,
    params={"topCatgryCd": tv_cat, "upCatgryCd": tv_cat})
log("위티TV", "GET", "/tv/favorites", r)

# 즐겨찾기 등록 - 실제 contsId 사용
fav_conts = real_tv_conts_id or real_conts_id or "test"
r = safe_request("POST", "/tv/favorites", headers=AUTH_JSON,
    json_body={"contsId": fav_conts, "favoritesYn": "Y"})
log("위티TV", "POST", "/tv/favorites", r)

r = safe_request("GET", "/tv/recommend", headers=AUTH, params={"catgryCd": tv_cat})
log("위티TV", "GET", "/tv/recommend", r)

r = safe_request("GET", "/tv/e-board", headers=AUTH)
log("위티TV", "GET", "/tv/e-board", r)

# ── 위티플레이 ──
print("\n=== 위티플레이 ===")

r = safe_request("GET", "/witti-play/game/main", headers=AUTH, params={"contsId": real_game_conts})
log("위티플레이", "GET", "/witti-play/game/main", r)

r = safe_request("GET", "/witti-play/game/score", headers=AUTH, params={"contsId": real_game_conts})
log("위티플레이", "GET", "/witti-play/game/score", r)

r = safe_request("POST", "/witti-play/game/score", headers=AUTH_JSON,
    json_body={"contsId": real_game_conts, "gameScore": 100, "gameHistSeq": 1})
log("위티플레이", "POST", "/witti-play/game/score", r)

r = safe_request("GET", "/witti-play/game/ranked", headers=AUTH, params={"contsId": real_game_conts})
log("위티플레이", "GET", "/witti-play/game/ranked", r)

r = safe_request("POST", "/witti-play/game/init", headers=AUTH_JSON)
log("위티플레이", "POST", "/witti-play/game/init", r)

# ── 위티박스(런처) ──
print("\n=== 위티박스(런처) ===")

r = safe_request("GET", "/config/apps/list", headers=AUTH_DEVICE,
    params={"ptnrId": "1102", "managedYn": "N"})
log("위티박스", "GET", "/config/apps/list", r)

r = safe_request("GET", "/witti-box/curriculum", headers=AUTH, params={"childId": child_id})
log("위티박스", "GET", "/witti-box/curriculum", r)

r = safe_request("POST", "/witti-app/attendance/curriculum", headers=AUTH_JSON,
    json_body={"isMidNight": "false"})
log("위티박스", "POST", "/witti-app/attendance/curriculum", r)

# report API들 - curriculumTp: int, childAge: string
curriculum_tp = 1 if target_age == "NEULBOM" else 0
r = safe_request("POST", "/report/parentReport", headers=AUTH_JSON,
    json_body={"curriculumTp": curriculum_tp, "childAge": target_age, "childId": child_id,
               "year": 2026, "month": 3, "week": 3, "reportType": "WEEK", "parentTp": "P"})
log("위티박스", "POST", "/report/parentReport", r)

r = safe_request("POST", "/report/teacherReport", headers=AUTH_JSON,
    json_body={"curriculumTp": curriculum_tp, "childAge": target_age, "childId": child_id,
               "year": 2026, "month": 3, "week": 3, "reportType": "WEEK"})
log("위티박스", "POST", "/report/teacherReport", r)

r = safe_request("POST", "/report/teacherActivityReport", headers=AUTH_JSON,
    json_body={"curriculumTp": curriculum_tp, "childAge": target_age, "childId": child_id,
               "year": 2026, "month": 3, "week": 3, "reportType": "WEEK"})
log("위티박스", "POST", "/report/teacherActivityReport", r)

r = safe_request("POST", "/mission/today/list", headers=AUTH_JSON,
    json_body={"childId": child_id, "ptnrSubTp": "1102"})
log("위티박스", "POST", "/mission/today/list", r)

# study/status - studyStatus 값: STUDY 또는 READY
r = safe_request("PATCH", "/authenticate/study/status", headers=AUTH_JSON_DEVICE,
    json_body={"studentId": student_id, "classId": class_id,
               "studyStatus": "STUDY", "loginId": USER_ID})
log("위티박스", "PATCH", "/authenticate/study/status", r)

# ── 위티App ──
print("\n=== 위티App ===")

r = safe_request("POST", "/witti-app/init/character", headers=AUTH_JSON)
log("위티App", "POST", "/witti-app/init/character", r)

# 캐릭터 생성 시도 (init 후 조회가 실패할 수 있으므로)
r = safe_request("POST", "/witti-app/character", headers=AUTH_JSON,
    json_body={"childId": child_id, "saveTp": "NEW", "nickName": "테스트봇",
               "wearingItemList": [], "invenItemList": []})
log("위티App", "POST", "/witti-app/character (생성)", r)

r = safe_request("GET", "/witti-app/character", headers=AUTH)
log("위티App", "GET", "/witti-app/character", r)

r = safe_request("POST", "/witti-app/init/mission", headers=AUTH_JSON)
log("위티App", "POST", "/witti-app/init/mission", r)

r = safe_request("POST", "/witti-app/init/attendance", headers=AUTH_JSON)
log("위티App", "POST", "/witti-app/init/attendance", r)

r = safe_request("POST", "/witti-app/promiseReward", headers=AUTH_JSON)
log("위티App", "POST", "/witti-app/promiseReward", r)

r = safe_request("POST", "/witti-app/promiseRewardInit", headers=AUTH_JSON)
log("위티App", "POST", "/witti-app/promiseRewardInit", r)

# instl-app-ver - X-device-info 헤더 추가
r = safe_request("GET", "/witti-app/instl-app-ver", headers=AUTH_DEVICE)
log("위티App", "GET", "/witti-app/instl-app-ver", r)

r = safe_request("GET", "/witti-app/child/list", headers=AUTH)
log("위티App", "GET", "/witti-app/child/list", r)

r = safe_request("GET", "/witti-app/mission", headers=AUTH, params={"misnGrp": "TUTORIAL"})
log("위티App", "GET", "/witti-app/mission", r)

r = safe_request("GET", "/witti-app/character/nickname-chk", headers=AUTH, params={"nickName": "테스트"})
log("위티App", "GET", "/witti-app/character/nickname-chk", r)

r = safe_request("GET", "/witti-app/character/nickname-random", headers=AUTH)
log("위티App", "GET", "/witti-app/character/nickname-random", r)

r = safe_request("GET", "/witti-app/main", headers=AUTH)
log("위티App", "GET", "/witti-app/main", r)

r = safe_request("GET", "/witti-app/reward/list", headers=AUTH, params={"rwrdClsTp": "P"})
log("위티App", "GET", "/witti-app/reward/list", r)

# reward/cheat - DEV에서만 동작
r = safe_request("POST", "/witti-app/reward/cheat", headers=AUTH_JSON)
log("위티App", "POST", "/witti-app/reward/cheat", r)

r = safe_request("GET", "/witti-app/live-sketch/animal", headers=AUTH)
log("위티App", "GET", "/witti-app/live-sketch/animal", r)

r = safe_request("POST", "/witti-app/init/live-sketch", headers=AUTH_JSON)
log("위티App", "POST", "/witti-app/init/live-sketch", r)

r = safe_request("GET", "/witti-app/shop-item/list", headers=AUTH)
log("위티App", "GET", "/witti-app/shop-item/list", r)

r = safe_request("GET", "/witti-app/shop-car/list", headers=AUTH)
log("위티App", "GET", "/witti-app/shop-car/list", r)

# badge - cornerTp 파라미터 추가
r = safe_request("GET", "/witti-app/badge", headers=AUTH, params={"cornerTp": "fire"})
log("위티App", "GET", "/witti-app/badge", r)

r = safe_request("POST", "/witti-app/badge", headers=AUTH_JSON,
    json_body={"cornerTp": "fire", "useYn": "Y"})
log("위티App", "POST", "/witti-app/badge", r)

r = safe_request("GET", "/witti-app/quiz", headers=AUTH,
    params={"cornerTp": "fire", "quizCatgry": "firefight"})
log("위티App", "GET", "/witti-app/quiz", r)

r = safe_request("GET", "/witti-app/quiz/gift", headers=AUTH, params={"cornerTp": "fire"})
log("위티App", "GET", "/witti-app/quiz/gift", r)

r = safe_request("POST", "/witti-app/quiz/gift", headers=AUTH_JSON,
    json_body={"badgeIdx": 2001})
log("위티App", "POST", "/witti-app/quiz/gift", r)

r = safe_request("GET", "/witti-app/zodiac/list", headers=AUTH)
log("위티App", "GET", "/witti-app/zodiac/list", r)

r = safe_request("POST", "/witti-app/zodiac/owned", headers=AUTH_JSON,
    params={"zodiacId": "1"})
log("위티App", "POST", "/witti-app/zodiac/owned", r)

r = safe_request("GET", "/witti-app/trip/area", headers=AUTH)
log("위티App", "GET", "/witti-app/trip/area", r)

r = safe_request("POST", "/witti-app/trip/area", headers=AUTH_JSON,
    json_body={"tripId": "1", "tripImg": "", "tripImgExt": "png"})
log("위티App", "POST", "/witti-app/trip/area", r)

r = safe_request("GET", "/witti-app/trip/area/quizReward", headers=AUTH, params={"areaCd": "0"})
log("위티App", "GET", "/witti-app/trip/area/quizReward", r)

r = safe_request("POST", "/witti-app/trip/area/quizReward", headers=AUTH_JSON,
    json_body={"areaCd": "0", "successCount": 3, "failCount": 0})
log("위티App", "POST", "/witti-app/trip/area/quizReward", r)

# ── 공통 ──
print("\n=== 공통 ===")

r = safe_request("GET", "/contents/random", headers=AUTH, params={"randomTp": "MSC"})
log("공통", "GET", "/contents/random", r)

# ── 학습 이력 등록 (IAS) ──
print("\n=== 학습 이력 등록 ===")

r = safe_request("POST", "/ias/school", headers=AUTH_JSON,
    json_body={"prodId": real_school_prod, "itemCd": "ITEM001", "curtniSeq": 1,
               "stdLtm": 60, "stdScore": 100, "stdTp": "C"})
log("학습이력", "POST", "/ias/school", r)

r = safe_request("POST", "/ias/school/ebook", headers=AUTH_JSON,
    json_body={"contsId": real_tv_conts_id or "EB001", "stdLtm": 60, "contsLtm": 120, "stdTp": "C", "latestPage": 5})
log("학습이력", "POST", "/ias/school/ebook", r)

r = safe_request("POST", "/ias/school/playground", headers=AUTH_JSON,
    json_body={"contsId": real_tv_conts_id or "PG001", "stdLtm": 60, "contsLtm": 120, "stdTp": "C"})
log("학습이력", "POST", "/ias/school/playground", r)

r = safe_request("POST", "/ias/tv", headers=AUTH_JSON,
    json_body={"contsId": real_tv_conts_id or "TV001", "stdLtm": 60, "contsLtm": 120, "stdTp": "C"})
log("학습이력", "POST", "/ias/tv", r)

r = safe_request("POST", "/ias/camp", headers=AUTH_JSON,
    json_body={"contsId": real_tv_conts_id or "CP001", "stdLtm": 60, "contsLtm": 120, "stdTp": "C"})
log("학습이력", "POST", "/ias/camp", r)

# ── 결과 요약 ──
pass_count = sum(1 for r in results if r[4] == "PASS")
fail_count = sum(1 for r in results if r[4] == "FAIL")
print(f"\n총 {len(results)}개 API | PASS: {pass_count} | FAIL: {fail_count}")

# ── 엑셀 저장 (기존 파일에 새 탭 추가) ──
date_str = datetime.now().strftime("%y%m%d")
file_path = f"C:/Users/User/Downloads/API_테스트결과_{date_str}.xlsx"
server_label = {"dev-api": "Dev", "qa-api": "QA", "api": "Prod"}.get(SERVER, SERVER)

if os.path.exists(file_path):
    wb = load_workbook(file_path)
    print(f"기존 파일 로드: {file_path}")
else:
    wb = Workbook()
    wb.remove(wb.active)
    print("새 파일 생성")

sheet_name = f"{server_label}_{date_str}"
# 동일 시트명 존재 시 삭제 후 재생성
if sheet_name in wb.sheetnames:
    del wb[sheet_name]
ws = wb.create_sheet(title=sheet_name)

# 스타일 정의
header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_font = Font(name="맑은 고딕", size=10)
cell_align = Alignment(vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center")
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
pass_font = Font(name="맑은 고딕", size=10, color="006100", bold=True)
fail_font = Font(name="맑은 고딕", size=10, color="9C0006", bold=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# 요약 행
ws.merge_cells("A1:I1")
summary_cell = ws["A1"]
summary_cell.value = (
    f"API 테스트 결과 — {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
    f"서버: {server_label}  |  계정: {USER_ID}  |  "
    f"총 {len(results)}개  |  PASS: {pass_count}  |  FAIL: {fail_count}"
)
summary_cell.font = Font(name="맑은 고딕", bold=True, size=12)
summary_cell.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 30

# 헤더: No. / 서버 / 카테고리 / Method / API Path / Status / 결과 / Result Data / 에러 상세
col_headers = ["No.", "서버", "카테고리", "Method", "API Path", "Status", "결과", "Result Data", "에러 상세"]
ws.append([])  # row 2 빈 줄
for col_idx, h in enumerate(col_headers, 1):
    cell = ws.cell(row=3, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
ws.row_dimensions[3].height = 25

# 데이터 행
for i, (tag, method, path, status, mark, result_data, error_detail) in enumerate(results, 1):
    row = i + 3
    values = [i, server_label, tag, method, path, status if status else "", mark, result_data, error_detail]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = cell_font
        cell.border = thin_border
        if col_idx in (1, 2, 4, 6, 7):
            cell.alignment = center_align
        else:
            cell.alignment = cell_align

    result_cell = ws.cell(row=row, column=7)
    if mark == "PASS":
        result_cell.fill = pass_fill
        result_cell.font = pass_font
    else:
        result_cell.fill = fail_fill
        result_cell.font = fail_font

# 카테고리별 배경색 교대
tag_colors = {}
color_toggle = ["F2F7FB", "FFFFFF"]
color_idx = 0
prev_tag = None
for i, (tag, *_) in enumerate(results):
    if tag != prev_tag:
        if tag not in tag_colors:
            tag_colors[tag] = color_toggle[color_idx % 2]
            color_idx += 1
        prev_tag = tag
    row = i + 4
    bg = PatternFill(start_color=tag_colors[tag], end_color=tag_colors[tag], fill_type="solid")
    for col in range(1, 8):  # Result Data, 에러 상세 제외
        c = ws.cell(row=row, column=col)
        if col != 7:  # 결과 열은 PASS/FAIL 색상 유지
            c.fill = bg

# 열 너비
ws.column_dimensions["A"].width = 6    # No.
ws.column_dimensions["B"].width = 9    # 서버
ws.column_dimensions["C"].width = 16   # 카테고리
ws.column_dimensions["D"].width = 9    # Method
ws.column_dimensions["E"].width = 48   # API Path
ws.column_dimensions["F"].width = 9    # Status
ws.column_dimensions["G"].width = 9    # 결과
ws.column_dimensions["H"].width = 60   # Result Data
ws.column_dimensions["I"].width = 60   # 에러 상세

# 자동 필터
ws.auto_filter.ref = f"A3:I{len(results) + 3}"

# 틀 고정
ws.freeze_panes = "A4"

wb.save(file_path)
print(f"엑셀 저장 완료: {file_path} (탭: {sheet_name})")
