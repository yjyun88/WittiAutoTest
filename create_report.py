import os
import time as pytime
import traceback
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

REPORT_HEADERS = [
    "test_time",
    "device",
    "server_env",
    "step",
    "class_name",
    "content_name",
    "content_thumb",
    "capture_screen",
    "is_video_playing",
    "status",
    "error_message",
    "duration_sec",
]

SUMMARY_SHEET_NAME = "Summary"
SAVE_EVERY_N_ROWS = 10
IMAGE_PX = 200

# Runtime caches to reduce repeated IO and image recompression
_REPORT_CACHE = {}
_PNG_OPT_CACHE = {}


def _header_style(cell):
    cell.fill = PatternFill(fill_type="solid", start_color="DDDDDD", end_color="DDDDDD")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(vertical="center", horizontal="center")


def _status_style(status):
    status_upper = (status or "").upper()
    if status_upper == "PASS":
        return Font(color="008000", bold=True), PatternFill(fill_type="solid", start_color="CCFFCC", end_color="CCFFCC")
    if status_upper == "FAIL":
        return Font(color="FF0000", bold=True), PatternFill(fill_type="solid", start_color="FFCCCC", end_color="FFCCCC")
    if status_upper == "SKIP":
        return Font(color="666666", bold=True), PatternFill(fill_type="solid", start_color="EEEEEE", end_color="EEEEEE")
    if status_upper == "ERROR":
        return Font(color="990000", bold=True), PatternFill(fill_type="solid", start_color="FFD9D9", end_color="FFD9D9")
    if status_upper == "RETRY":
        return Font(color="CC7A00", bold=True), PatternFill(fill_type="solid", start_color="FFE8CC", end_color="FFE8CC")
    return Font(), PatternFill(fill_type=None)


def _normalize_server_api(server_env):
    mapping = {
        "Prod": "api",
        "QA": "qa-api",
        "Dev": "dev-api",
        "Total-Test": "total-test-api",
    }
    return mapping.get(server_env, server_env)


def _infer_step(class_name, content_name):
    cn = (class_name or "").upper()
    cont = (content_name or "").upper()
    if "MEW" in cn or "MEW" in cont:
        return "MEW"
    if "KOR" in cn or "MTH" in cn or "SCI" in cn:
        return "ARAM_WORLD"
    return "WITTI_BOX"


def _get_header_map(ws):
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    return {str(cell.value): cell.column for cell in header_row if cell.value}


def _format_duration_hms(total_seconds):
    try:
        total_seconds = int(round(float(total_seconds or 0)))
    except Exception:
        total_seconds = 0
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    return f"{hours}시간 {minutes}분 {seconds}초"


def _quote_sheet_title(title):
    return "'" + str(title).replace("'", "''") + "'"


def _duration_text_formula(total_seconds_expr):
    expr = f"({total_seconds_expr})"
    return (
        f'=INT({expr}/3600)&"시간 "&'
        f'INT(MOD({expr},3600)/60)&"분 "&'
        f'ROUND(MOD({expr},60),0)&"초"'
    )


def _ensure_headers(ws):
    # Repair legacy sheet where headers were shifted by one column (A is empty, B starts with test_time).
    if (
        ws.max_column >= 2
        and ws.cell(row=1, column=1).value is None
        and ws.cell(row=1, column=2).value in REPORT_HEADERS
    ):
        first_col_has_values = any(ws.cell(row=r, column=1).value is not None for r in range(1, ws.max_row + 1))
        if not first_col_has_values:
            ws.delete_cols(1, 1)

    # Brand-new empty sheet: write headers from column A explicitly.
    if ws.max_row == 1 and all(ws.cell(row=1, column=c).value is None for c in range(1, ws.max_column + 1)):
        for idx, header in enumerate(REPORT_HEADERS, start=1):
            ws.cell(row=1, column=idx, value=header)

    current = _get_header_map(ws)
    if "server_api" in current:
        ws.delete_cols(current["server_api"], 1)
        current = _get_header_map(ws)

    next_col = ws.max_column + 1
    for header in REPORT_HEADERS:
        if header not in current:
            ws.cell(row=1, column=next_col, value=header)
            current[header] = next_col
            next_col += 1

    # Re-style full header row
    for col in range(1, ws.max_column + 1):
        _header_style(ws.cell(row=1, column=col))

    ws.freeze_panes = "A2"
    return current


def optimize_png(
    input_path,
    output_path=None,
    compress_level=9,
    max_colors=128,
    max_width=200,
    max_height=200,
):
    """Optimize PNG once per file-state and return optimized path."""
    if not input_path or not os.path.exists(input_path):
        return None

    if output_path is None:
        output_path = input_path

    try:
        st = os.stat(input_path)
        cache_key = (
            os.path.abspath(input_path),
            st.st_mtime_ns,
            st.st_size,
            compress_level,
            max_colors,
            max_width,
            max_height,
            os.path.abspath(output_path),
        )
        cached = _PNG_OPT_CACHE.get(cache_key)
        if cached and os.path.exists(cached):
            return cached

        with PILImage.open(input_path) as img:
            img.thumbnail((max_width, max_height), PILImage.LANCZOS)
            img = img.convert("P", palette=PILImage.ADAPTIVE, colors=max_colors)
            img.save(
                output_path,
                format="PNG",
                optimize=True,
                compress_level=compress_level,
            )

        _PNG_OPT_CACHE[cache_key] = output_path
        return output_path
    except Exception as e:
        print(f"[WARN] optimize_png failed for {input_path}: {e}")
        return input_path


def adjust_column_widths(ws, padding=2, scale=1.15, max_width=60):
    """Adjust text column widths. Image columns are handled separately."""
    header = _ensure_headers(ws)
    image_cols = {header.get("content_thumb"), header.get("capture_screen")}

    for col in range(1, ws.max_column + 1):
        if col in image_cols:
            continue

        col_letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))

        width = min(max_width, (max_len + padding) * scale)
        ws.column_dimensions[col_letter].width = max(width, 10)


def adjust_row_heights(ws, row, px_size=IMAGE_PX):
    """Set image columns width and target row height for a report row."""
    pixel_to_pt = 72 / 96
    pixel_to_char = 1 / 7

    header = _ensure_headers(ws)
    cap_idx = header.get("capture_screen")
    thumb_idx = header.get("content_thumb")

    if cap_idx is None:
        raise ValueError("Header 'capture_screen' not found.")
    if thumb_idx is None:
        raise ValueError("Header 'content_thumb' not found.")

    cap_letter = get_column_letter(cap_idx)
    thumb_letter = get_column_letter(thumb_idx)

    ws.column_dimensions[cap_letter].width = px_size * pixel_to_char
    ws.column_dimensions[thumb_letter].width = px_size * pixel_to_char
    ws.row_dimensions[row].height = px_size * pixel_to_pt


def apply_borders_to_filled_rows(ws, only_row=None):
    thin = Side(border_style="thin", color="000000")
    bd = Border(top=thin, bottom=thin, left=thin, right=thin)

    if only_row is not None:
        for c in range(1, ws.max_column + 1):
            ws.cell(row=only_row, column=c).border = bd
        return

    for r in range(2, ws.max_row + 1):
        if any(ws.cell(row=r, column=c).value is not None for c in range(1, ws.max_column + 1)):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).border = bd


def _write_summary_sheet(wb, ws):
    summary_ws = wb[SUMMARY_SHEET_NAME] if SUMMARY_SHEET_NAME in wb.sheetnames else wb.create_sheet(SUMMARY_SHEET_NAME)
    summary_ws.delete_rows(1, summary_ws.max_row or 1)
    if wb.sheetnames[0] != SUMMARY_SHEET_NAME:
        wb._sheets.remove(summary_ws)
        wb._sheets.insert(0, summary_ws)

    summary_header = ["date", "total", "pass", "fail", "skip", "error", "retry", "other", "pass_rate", "duration_total"]
    summary_ws.append(summary_header)
    for cell in summary_ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    duration_seconds_terms = []

    for sheet in wb.worksheets:
        if sheet.title == SUMMARY_SHEET_NAME:
            continue
        header = _ensure_headers(sheet)
        status_col = header.get("status") or header.get("is_video_playing")
        content_col = header.get("content_name")
        duration_col = header.get("duration_sec")
        status_letter = get_column_letter(status_col) if status_col else None
        content_letter = get_column_letter(content_col) if content_col else None
        quoted_title = _quote_sheet_title(sheet.title)
        total_formula = (
            f'=COUNTA({quoted_title}!{content_letter}:{content_letter})-1'
            if content_letter
            else "0"
        )
        pass_formula = f'=COUNTIF({quoted_title}!{status_letter}:{status_letter},"PASS")' if status_letter else "0"
        fail_formula = f'=COUNTIF({quoted_title}!{status_letter}:{status_letter},"FAIL")' if status_letter else "0"
        skip_formula = f'=COUNTIF({quoted_title}!{status_letter}:{status_letter},"SKIP")' if status_letter else "0"
        error_formula = f'=COUNTIF({quoted_title}!{status_letter}:{status_letter},"ERROR")' if status_letter else "0"
        retry_formula = f'=COUNTIF({quoted_title}!{status_letter}:{status_letter},"RETRY")' if status_letter else "0"
        other_formula = (
            f"=B{summary_ws.max_row + 1}-SUM(C{summary_ws.max_row + 1}:G{summary_ws.max_row + 1})"
        )
        pass_rate_formula = (
            f'=IF(B{summary_ws.max_row + 1}>0,C{summary_ws.max_row + 1}/B{summary_ws.max_row + 1},0)'
        )
        duration_formula = (
            _duration_text_formula(
                f"SUM({quoted_title}!{get_column_letter(duration_col)}:{get_column_letter(duration_col)})"
            )
            if duration_col
            else "=\"0시간 0분 0초\""
        )
        if duration_col:
            duration_seconds_terms.append(
                f"SUM({quoted_title}!{get_column_letter(duration_col)}:{get_column_letter(duration_col)})"
            )
        summary_ws.append([
            sheet.title,
            total_formula,
            pass_formula,
            fail_formula,
            skip_formula,
            error_formula,
            retry_formula,
            other_formula,
            pass_rate_formula,
            duration_formula,
        ])

    last_data_row = summary_ws.max_row
    summary_ws.append([])
    first_data_row = 2
    month_total_row = last_data_row + 2
    total_range = f"B{first_data_row}:B{last_data_row}" if last_data_row >= first_data_row else None
    pass_range = f"C{first_data_row}:C{last_data_row}" if last_data_row >= first_data_row else None
    fail_range = f"D{first_data_row}:D{last_data_row}" if last_data_row >= first_data_row else None
    skip_range = f"E{first_data_row}:E{last_data_row}" if last_data_row >= first_data_row else None
    error_range = f"F{first_data_row}:F{last_data_row}" if last_data_row >= first_data_row else None
    retry_range = f"G{first_data_row}:G{last_data_row}" if last_data_row >= first_data_row else None
    other_range = f"H{first_data_row}:H{last_data_row}" if last_data_row >= first_data_row else None
    month_duration_formula = (
        _duration_text_formula("+".join(duration_seconds_terms))
        if duration_seconds_terms
        else "=\"0시간 0분 0초\""
    )
    summary_ws.append([
        "MONTH_TOTAL",
        f"=SUM({total_range})" if total_range else 0,
        f"=SUM({pass_range})" if pass_range else 0,
        f"=SUM({fail_range})" if fail_range else 0,
        f"=SUM({skip_range})" if skip_range else 0,
        f"=SUM({error_range})" if error_range else 0,
        f"=SUM({retry_range})" if retry_range else 0,
        f"=SUM({other_range})" if other_range else 0,
        f"=IF(B{month_total_row}>0,C{month_total_row}/B{month_total_row},0)",
        month_duration_formula,
    ])
    for cell in summary_ws[summary_ws.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(2, summary_ws.max_row + 1):
        summary_ws.cell(row=row, column=9).number_format = "0.00%"

    for row in range(1, summary_ws.max_row + 1):
        for col in range(1, summary_ws.max_column + 1):
            summary_ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # Column widths based on displayed content (not formula length)
    col_widths = {
        "date": 16,
        "total": 8,
        "pass": 8,
        "fail": 8,
        "skip": 8,
        "error": 8,
        "retry": 8,
        "other": 8,
        "pass_rate": 12,
        "duration_total": 22,
    }
    for col_idx, header_name in enumerate(summary_header, start=1):
        summary_ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header_name, 12)


def _flush_report(file_path):
    ctx = _REPORT_CACHE.get(file_path)
    if not ctx:
        return

    wb = ctx["wb"]
    ws = ctx["ws"]

    # openpyxl can fail to save when an image ref points to a closed stream.
    # Freeze image payloads to in-memory bytes before writing the workbook.
    for sheet in wb.worksheets:
        images = list(getattr(sheet, "_images", []) or [])
        kept = []
        for img in images:
            try:
                payload = img._data()
                if not isinstance(payload, (bytes, bytearray)) or len(payload) == 0:
                    continue
                payload = bytes(payload)
                img.ref = BytesIO(payload)
                img._data = (lambda imgb=payload: imgb)
                kept.append(img)
            except Exception as e:
                print(f"[WARN] dropped broken image before save: {e}")
        sheet._images = kept

    adjust_column_widths(ws, padding=2, scale=1.15)
    _write_summary_sheet(wb, ws)
    if getattr(wb, "calculation", None) is not None:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    wb.save(file_path)
    ctx["pending_rows"] = 0


def flush_all_reports():
    for file_path in list(_REPORT_CACHE.keys()):
        try:
            _flush_report(file_path)
        except Exception as e:
            # During interpreter shutdown, workbook/image internals may already be closed.
            if "closed file" in str(e).lower():
                continue
            print(f"[WARN] Failed to flush report {file_path}: {e}")



def create_report():
    """
    Create/load monthly workbook and return today's sheet.
    """
    today = datetime.now()
    folder_name = today.strftime("%Y.%m")
    sheet_name = today.strftime("%y.%m.%d")

    base_dir = os.getcwd()
    target_dir = os.path.join(base_dir, "test_report")
    os.makedirs(target_dir, exist_ok=True)

    file_name = f"REPORT_{folder_name}.xlsx"
    file_path = os.path.join(target_dir, file_name)

    ctx = _REPORT_CACHE.get(file_path)
    if ctx and os.path.exists(file_path):
        wb = ctx["wb"]
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            _ensure_headers(ws)
        ctx["ws"] = ws
        return file_path, wb, ws

    if os.path.exists(file_path):
        try:
            wb = load_workbook(file_path)
        except Exception:
            corrupt_name = f"{file_path}.corrupt_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            try:
                os.replace(file_path, corrupt_name)
                print(f"[WARN] Corrupted report detected. Backed up to: {corrupt_name}")
            except Exception as backup_error:
                print(f"[WARN] Failed to backup corrupted report: {backup_error}")
            wb = Workbook()
    else:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        if wb.sheetnames == ["Sheet"] and wb["Sheet"].max_row == 1 and wb["Sheet"]["A1"].value is None:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

    _ensure_headers(ws)

    _REPORT_CACHE[file_path] = {"wb": wb, "ws": ws, "pending_rows": 0}
    return file_path, wb, ws


def input_excel(
    video_playing,
    class_name,
    content_name,
    file_path,
    wb,
    ws,
    capture_path,
    thumb_path,
    width=IMAGE_PX,
    height=IMAGE_PX,
    padding_chars=2,
    test_time=None,
    device=None,
    server_env=None,
    server_api=None,
    step=None,
    error_message=None,
    duration_sec=None,
    auto_save=True,
):
    """
    Append one test row into report worksheet with optional images and metadata.
    """
    try:
        header = _ensure_headers(ws)

        row = ws.max_row + 1
        status = (video_playing or "UNKNOWN").upper()

        test_time = test_time or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        server_env = server_env or os.environ.get("REPORT_SERVER_ENV", "")
        step = step or os.environ.get("REPORT_STEP", "") or _infer_step(class_name, content_name)
        error_message = (error_message or "")[:300]

        row_values = {
            "test_time": test_time,
            "device": device or os.environ.get("REPORT_DEVICE", ""),
            "server_env": server_env,
            "step": step,
            "class_name": class_name,
            "content_name": content_name,
            "is_video_playing": status,
            "status": status,
            "error_message": error_message,
            "duration_sec": duration_sec,
        }

        for key, value in row_values.items():
            col = header[key]
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(vertical="center", horizontal="center")

        # Attach thumbnail image
        thumb_col = header["content_thumb"]
        if thumb_path:
            thumb_opt = optimize_png(thumb_path, compress_level=9, max_width=width, max_height=height)
            if thumb_opt and os.path.exists(thumb_opt):
                with open(thumb_opt, "rb") as f:
                    thumb_bytes = f.read()
                thumb_buf = BytesIO(thumb_bytes)
                thumb_img = XLImage(thumb_buf)
                thumb_img.width, thumb_img.height = width, height
                thumb_img._data = lambda imgb=thumb_bytes: imgb
                ws.add_image(thumb_img, f"{get_column_letter(thumb_col)}{row}")

        # Attach capture image
        cap_col = header["capture_screen"]
        cap_opt = optimize_png(capture_path, compress_level=9, max_width=width, max_height=height)
        if cap_opt and os.path.exists(cap_opt):
            with open(cap_opt, "rb") as f:
                cap_bytes = f.read()
            cap_buf = BytesIO(cap_bytes)
            cap_img = XLImage(cap_buf)
            cap_img.width, cap_img.height = width, height
            cap_img._data = lambda imgb=cap_bytes: imgb
            ws.add_image(cap_img, f"{get_column_letter(cap_col)}{row}")

        # Style status-related cells
        status_font, status_fill = _status_style(status)

        status_cell = ws.cell(row=row, column=header["status"])
        status_cell.font = status_font
        status_cell.fill = status_fill

        video_status_col = header.get("is_video_playing")
        if video_status_col:
            video_status_cell = ws.cell(row=row, column=video_status_col)
            video_status_cell.font = status_font
            video_status_cell.fill = status_fill
            video_status_cell.alignment = Alignment(vertical="center", horizontal="center")

        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).alignment = Alignment(vertical="center", horizontal="center")

        # Layout and border for this row
        adjust_row_heights(ws, row, px_size=max(width, height))
        apply_borders_to_filled_rows(ws, only_row=row)

        # Save policy: periodic + immediate on FAIL/ERROR
        ctx = _REPORT_CACHE.setdefault(file_path, {"wb": wb, "ws": ws, "pending_rows": 0})
        ctx["wb"] = wb
        ctx["ws"] = ws
        ctx["pending_rows"] = ctx.get("pending_rows", 0) + 1

        # Default to immediate save to keep report behavior predictable.
        should_flush = bool(auto_save)

        if should_flush:
            _flush_report(file_path)

        print(
            f"Row {row} appended: class='{class_name}', content='{content_name}', "
            f"status='{status}', step='{step}'"
        )
        return True
    except Exception as e:
        print("=== Report row append failed ===")
        traceback.print_exc()
        print(f"[WARN] input_excel continue on error: {e}")
        return False


def report_thumbnail_error(img_path, child_nm, image_folder_abs, message, started_at):
    from box_ACT import capture_screen
    capture_path, base = capture_screen(img_path, child_nm)
    file_path, wb, ws = create_report()
    thumb_path = os.path.join(image_folder_abs, os.path.basename(img_path))
    input_excel(
        "ERROR",
        child_nm,
        base,
        file_path,
        wb,
        ws,
        capture_path,
        thumb_path,
        error_message=message,
        duration_sec=round(pytime.perf_counter() - started_at, 2),
    )
